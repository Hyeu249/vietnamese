# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from . import CONST
from ...utilities.models import CONST as UTILITIES_CONST

from odoo.exceptions import ValidationError
from datetime import timedelta
from .common_utils import generate_token, format_field_date
import logging
import io
import os
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import tempfile
from odoo.exceptions import UserError
from odoo import _, api, fields, models
from datetime import datetime, timedelta
from openpyxl.styles import Border, Side, Alignment


class ProposedLiquidation(models.Model):
    _name = "ship.proposed.liquidation"
    _description = "Đề nghị thanh lý"
    _inherit = ["utilities.approval.status"]
    _check_company_auto = True

    name = fields.Char("Name", tracking=True)
    name_decision = fields.Char("Name Decision", tracking=True)
    name_agreement = fields.Char("Name Agreement", tracking=True)
    description = fields.Char("Description", tracking=True)
    replace_date_from = fields.Date("Replace date from", tracking=True)
    replace_date_to = fields.Date("Replace date to", tracking=True)
    proposed_date = fields.Date("Proposed Date", tracking=True)
    liquidation_method = fields.Char("Liquidation Method", tracking=True)
    liquidation_location = fields.Char("Liquidation Location", tracking=True)
    liquidation_agency = fields.Char("Liquidation Agency", tracking=True)

    # related
    name_for_noti = fields.Char(
        related="name",
        string="Name",
    )

    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    # relations
    replacement_diary_ids = fields.One2many(
        "ship.replacement.diary",
        "proposed_liquidation_id",
        string="Replacement diary",
        tracking=True,
    )

    # sequential
    ref = fields.Char(string="Code", default=lambda self: _("New"))

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            model_name = "ship.proposed.liquidation"
            vals["ref"] = self.env["ir.sequence"].next_by_code(model_name)
        result = super(ProposedLiquidation, self).create(vals_list)

        return result

    def write(self, vals):
        self.ensure_one()
        result = super(ProposedLiquidation, self).write(vals)

        if "approval_status" in vals:
            if self._is_approved():
                self.suggested_to_discard_to_all_entities()
            elif self._is_rejected():
                self.not_propose_to_replace_material_entities()

        return result

    def name_get(self):
        result = []
        for report in self:
            name = report.ref or _("New")
            result.append((report.id, name))
        return result

    def get_materials_by_replace_date(self):
        self.ensure_one()
        records = self.env["ship.replacement.diary"].search_read(
            domain=[
                ("date", ">=", self.replace_date_from),
                ("date", "<=", self.replace_date_to),
            ],
            fields=["id"],  # Specify the fields to read, in this case, only 'id'
        )

        replacement_diary_ids = [record["id"] for record in records]

        if replacement_diary_ids:
            self.replacement_diary_ids = [(6, 0, replacement_diary_ids)]
        else:
            self.replacement_diary_ids = False

    def create_liquidation_minute(self):
        self.ensure_one()
        if not self.liquidation_minute_ids:
            self.env["ship.liquidation.minute"].create(
                {
                    "proposed_liquidation_id": self.id,
                }
            )

    def custom_export_proposal_form_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = "/mnt/extra-addons/report_template/dntl.xlsx"  # Replace with the actual path
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###
        align_center = Alignment(horizontal="center", vertical="center")
        side_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
        )
        ###Đề nghị số
        worksheet.cell(row=2, column=1, value=f"Đề nghị số: {self.name}")
        worksheet.merge_cells("A2:K2")
        worksheet[f"A2"].alignment = align_center
        ### Tên tàu
        worksheet.cell(row=7, column=1, value=f"{self.company_id.name}")
        worksheet.merge_cells("A7:B7")
        worksheet[f"A7"].alignment = align_center
        ### Ngày thanh lý
        worksheet.cell(row=7, column=5, value=f"{self.proposed_date}")
        worksheet.merge_cells("E7:F7")
        worksheet[f"E7"].alignment = align_center
        ##Hình thức thanh lý
        worksheet.cell(row=7, column=7, value=f"{self.liquidation_method}")
        worksheet[f"A2"].alignment = align_center
        ###Vị trí thanh lý
        worksheet.cell(row=7, column=8, value=f"{self.liquidation_location}")
        worksheet.merge_cells("H7:J7")
        worksheet[f"H7"].alignment = align_center
        ## Đơn vị tiếp nhận
        worksheet.cell(row=7, column=11, value=f"{self.liquidation_agency}")

        row = 11
        index = 0
        for replacement_diary_id in self.replacement_diary_ids:
            worksheet.insert_rows(row)
            index += 1
            worksheet.cell(row=row, column=1, value=index)
            if replacement_diary_id.material_entity_id:
                worksheet.cell(
                    row=row,
                    column=2,
                    value=f"{replacement_diary_id.material_id.name} (Vật tư thực tế {replacement_diary_id.material_entity_id.ref})",
                )
            else:
                worksheet.cell(
                    row=row,
                    column=2,
                    value=replacement_diary_id.material_id.name,
                )
            worksheet.cell(
                row=row,
                column=3,
                value=replacement_diary_id.material_description,
            )
            worksheet.cell(
                row=row,
                column=4,
                value=replacement_diary_id.internal_code,
            )
            worksheet.cell(
                row=row, column=5, value=replacement_diary_id.material_id.unit
            )
            worksheet.cell(
                row=row,
                column=6,
                value=replacement_diary_id.quantity,
            )
            worksheet.cell(row=row, column=7, value=replacement_diary_id.reason)
            worksheet.cell(row=row, column=8, value=replacement_diary_id.date)
            worksheet.cell(
                row=row,
                column=9,
                value=(replacement_diary_id.year_used),
            )
            worksheet.cell(row=row, column=10, value=replacement_diary_id.condition)
            for col in range(1, 12):  # Column A to K (1 to 11 in 1-indexed system)
                cell = worksheet.cell(row=row, column=col)
                cell.border = side_border
                cell.alignment = align_center

            row += 1
        ##alingment

        worksheet.merge_cells(f"A{row+10}:D{row+10}")
        worksheet[f"A{row+10}"].alignment = align_center

        worksheet.merge_cells(f"H{row+10}:K{row+10}")
        worksheet[f"H{row+10}"].alignment = align_center

        worksheet.merge_cells(f"E{row+11}:G{row+11}")
        worksheet[f"E{row+11}"].alignment = align_center

        worksheet.merge_cells(f"A{row+11}:D{row+11}")
        worksheet[f"A{row+11}"].alignment = align_center

        worksheet.merge_cells(f"H{row+11}:K{row+11}")
        worksheet[f"H{row+11}"].alignment = align_center

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Đề nghị thanh lý số {self.name}.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action

    def custom_export_decision_form_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = "/mnt/extra-addons/report_template/qdtl.xlsx"  # Replace with the actual path
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###
        align_center = Alignment(horizontal="center", vertical="center")
        side_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
        )
        date_object = fields.Date.from_string(self.proposed_date)
        day = date_object.day
        month = date_object.month
        year = date_object.year
        ###Số
        worksheet.cell(row=5, column=1, value=f"Số: {self.name_decision}")
        worksheet.merge_cells("A5:D5")
        worksheet[f"A5"].alignment = align_center
        ### Điều 1
        worksheet.cell(
            row=9,
            column=1,
            value=f" Điều 1: Thanh lý các phụ tùng, vật tư trong bản đề nghị số: {self.name}, ngày {day} tháng {month} năm {year} của {self.company_id.name} ",
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Quyết định thanh lý số {self.name_decision}.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action

    def custom_export_agreement_form_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = "/mnt/extra-addons/report_template/bbtl.xlsx"  # Replace with the actual path
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###

        align_center = Alignment(horizontal="center", vertical="center")
        side_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
        )
        date_object = fields.Date.from_string(self.proposed_date)
        day = date_object.day
        month = date_object.month
        year = date_object.year
        ###Biên bản số
        worksheet.cell(row=3, column=1, value=f"Biên bản số: {self.name_agreement}")
        worksheet.merge_cells("A3:K3")
        worksheet[f"A3"].alignment = align_center
        ### Căn cứ
        worksheet.cell(
            row=5,
            column=1,
            value=f"  Căn cứ vào quyết định thanh lý số: {self.name_decision}  Ngày {day} tháng {month} năm {year}, tàu tiến hành thanh lý phụ tùng vật tư với nội dung như sau:",
        )

        ### Tên tàu
        worksheet.cell(row=7, column=1, value=f"{self.company_id.name}")
        worksheet.merge_cells("A7:B7")
        worksheet[f"A7"].alignment = align_center
        ### Ngày thanh lý
        worksheet.cell(row=7, column=5, value=f"{self.proposed_date}")
        worksheet.merge_cells("E7:F7")
        worksheet[f"E7"].alignment = align_center
        ##Hình thức thanh lý
        worksheet.cell(row=7, column=7, value=f"{self.liquidation_method}")
        worksheet[f"A2"].alignment = align_center
        ###Vị trí thanh lý
        worksheet.cell(row=7, column=8, value=f"{self.liquidation_location}")
        worksheet.merge_cells("H7:J7")
        worksheet[f"H7"].alignment = align_center
        ## Đơn vị tiếp nhận
        worksheet.cell(row=7, column=11, value=f"{self.liquidation_agency}")

        row = 11
        index = 0
        for replacement_diary_id in self.replacement_diary_ids:
            worksheet.insert_rows(row)
            index += 1
            worksheet.cell(row=row, column=1, value=index)
            if replacement_diary_id.material_entity_id:
                worksheet.cell(
                    row=row,
                    column=2,
                    value=f"{replacement_diary_id.material_id.name} (Vật tư thực tế {replacement_diary_id.material_entity_id.ref})",
                )
            else:
                worksheet.cell(
                    row=row,
                    column=2,
                    value=replacement_diary_id.material_id.name,
                )
            worksheet.cell(
                row=row,
                column=3,
                value=replacement_diary_id.material_description,
            )
            worksheet.cell(
                row=row,
                column=4,
                value=replacement_diary_id.internal_code,
            )
            worksheet.cell(
                row=row, column=5, value=replacement_diary_id.material_id.unit
            )
            worksheet.cell(
                row=row,
                column=6,
                value=replacement_diary_id.quantity,
            )
            worksheet.cell(row=row, column=7, value=replacement_diary_id.reason)
            worksheet.cell(row=row, column=8, value=replacement_diary_id.date)
            worksheet.cell(
                row=row,
                column=9,
                value=(replacement_diary_id.year_used),
            )
            worksheet.cell(row=row, column=10, value=replacement_diary_id.condition)
            for col in range(1, 12):  # Column A to K (1 to 11 in 1-indexed system)
                cell = worksheet.cell(row=row, column=col)
                cell.border = side_border
                cell.alignment = align_center

            row += 1
        ##alingment

        worksheet.merge_cells(f"E{row+12}:G{row+12}")
        worksheet[f"E{row+12}"].alignment = align_center

        worksheet.merge_cells(f"A{row+12}:D{row+12}")
        worksheet[f"A{row+12}"].alignment = align_center

        worksheet.merge_cells(f"H{row+12}:K{row+12}")
        worksheet[f"H{row+12}"].alignment = align_center

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Biên bản thanh lý số {self.name_agreement}.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action

    # def action_approve(self):
    #     self.ensure_one()
    #     super(ProposedLiquidation, self).action_approve()
    #     # suggest to discard the corresponding material entities
    #     for replacement_diary in self.replacement_diary_ids:
    #         replacement_diary.material_entity_id.write(
    #             {
    #                 "is_currently_proposed_to_replace": False,
    #                 "is_suggested_to_discard": True,
    #             }
    #         )

    def suggested_to_discard_to_all_entities(self):
        self.ensure_one()
        for replacement_diary in self.replacement_diary_ids:
            replacement_diary.material_entity_id.write(
                {
                    "is_currently_proposed_to_replace": False,
                    "is_suggested_to_discard": True,
                }
            )

    # def action_reject(self):
    #     self.ensure_one()
    #     super(ProposedLiquidation, self).action_reject()
    #     # set is_currently_proposed_to_replace to False for the corresponding material entities,
    #     # and do not discard them
    #     for replacement_diary in self.replacement_diary_ids:
    #         replacement_diary.material_entity_id.write(
    #             {
    #                 "is_currently_proposed_to_replace": False,
    #             }
    #         )

    def not_propose_to_replace_material_entities(self):
        self.ensure_one()
        for replacement_diary in self.replacement_diary_ids:
            entity = replacement_diary.material_entity_id
            entity.is_currently_proposed_to_replace = False
