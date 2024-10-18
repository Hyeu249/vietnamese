import io
import os
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import tempfile
from odoo.exceptions import UserError
from odoo import api, fields, models
from datetime import datetime, timedelta  # Import timedelta



class JobQuoteReportBatch(models.Model):
    _name = "ship.job.quote.report.batch"
    _description = "Job Quote Report Batch"
    _check_company_auto = True

    name = fields.Char("Name")
    date_start = fields.Date("Date From", required=True)
    date_end = fields.Date("Date To", required=True)

    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    # relations
    line_ids = fields.One2many(
        "ship.job.quote.report", "batch_id", string="Report Lines"
    )

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get("name"):
                model_code = "ship.job.quote.report.batch"
                vals["name"] = self.env["ir.sequence"].next_by_code(model_code) or _('New')
        return super(JobQuoteReportBatch, self).create(vals_list)

    def generate_report(self):
        self.ensure_one()
        self.env["ship.job.quote.report"].search(
            [("batch_id", "=", self.id)]
        ).unlink()
        self.env.cr.execute(
                """
                SELECT a.finished_at, a.approved_implement_date, j.name AS maintenance_scope_name, a.note, b.ship_location, b.problem_description,  case when 
                j.maintenance_type = 'THRESHOLD' then 'Định kỳ' 
                when  j.maintenance_type = 'ARISE' then 'Phát sinh'
                when  j.maintenance_type = 'CONSUMPTION' then 'Định kỳ' end as reason, 
                a.implement_date, b.week_number, c.name AS job_name, 
                CASE 
                    WHEN k.name IS NOT NULL THEN k.name
                    ELSE 'Thuyền viên'
                END AS suppplier_name, 
                g.labor_cost, e.name AS material_name, e.origin AS material_origin, h.average_quote_price, f.total_hours, j.last_maintenance_date
                FROM ship_job_quote a
                LEFT JOIN ship_maintenance_scope_report b ON a.maintenance_scope_report_id=b.id 
                LEFT JOIN ship_job c ON a.job_id = c.id
                LEFT JOIN ship_material_assignment d ON a.id=d.job_quote_id
                LEFT JOIN ship_material e ON d.material_id= e.id 
                LEFT JOIN ship_material_entity f ON d.material_entity_id=f.id 
                LEFT JOIN ship_job_supplier_quote g ON a.job_supplier_quote_id=g.id 
                LEFT JOIN ship_material_quote h ON d.material_id=h.material_id
                LEFT JOIN ship_maintenance_scope j ON b.maintenance_scope_id=j.id 
                LEFT JOIN ship_supplier k ON g.supplier_id=k.id
                WHERE (a.finished_at BETWEEN %s AND %s) OR (a.approved_implement_date BETWEEN %s AND %s + interval '7 days') and j.name is not null
                ORDER BY j.name
                """,
                (self.date_start, self.date_end, self.date_end, self.date_end),
            )
        results = self.env.cr.fetchall()

        data = [{
            'finished_date': result[0],
            'approved_implement_date': result[1],
            'maintenance_scope_name': result[2],
            'note': result[3],
            'ship_location': result[4],
            'problem_description': result[5],
            'reason': result[6],
            'implement_date': result[7],
            'week_number': result[8],
            'job_name': result[9],
            'supplier_name': result[10],
            'labor_cost': result[11],
            'material_name': result[12],
            'material_origin': result[13],
            'material_average_quote_price': result[14],
            'total_hours': result[15],
            'last_implement_date': result[16],
            'batch_id': self.id,
        } for result in results]

        self.env["ship.job.quote.report"].create(data)


        # Function to adjust merged cells after inserting rows

    def custom_export_to_xlsx(self):
    # Function to adjust merged cells after inserting rows


        # Load the custom XLSX template
        template_path = '/mnt/extra-addons/report_template/bgt.xlsx'  # Replace with the actual path
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")
        

        week_number = self.date_start.isocalendar()[1]

        worksheet['A1']= f"BÁO CÁO GIAO BAN TUẦN {week_number}"
        worksheet['D4']= f"{week_number}"
        worksheet['D605']= f"{week_number+1}"

        # Add data to the XLSX file
        data_row_9 = []  # List to store data for rows starting from 9
        data_row_100 = []  # List to store data for rows starting from 100

        for report in self.line_ids:
            if report.approved_implement_date and self.date_end <= report.approved_implement_date <= (self.date_end + timedelta(days=7)):
                data_row_100.append(report)
            else:
                data_row_9.append(report)

        # Start pasting data from row 9
        row = 9
        for report in data_row_9:
            # if row >= 9:
                # Insert a new row if the current row is 78
                # worksheet.insert_rows(row)
                # adjust_merged_cells(worksheet, row, 1)
            worksheet.cell(row=row, column=3, value=report.maintenance_scope_name)
            worksheet.cell(row=row, column=5, value=report.problem_description)
            worksheet.cell(row=row, column=6, value=report.ship_location)
            worksheet.cell(row=row, column=9, value=report.reason)
            worksheet.cell(row=row, column=10, value=report.finished_date)
            worksheet.cell(row=row, column=11, value=report.job_name)
            worksheet.cell(row=row, column=12, value=report.supplier_name)
            worksheet.cell(row=row, column=13, value=report.labor_cost)
            worksheet.cell(row=row, column=14, value=report.material_average_quote_price)
            worksheet.cell(row=row, column=15, value=report.material_name)
            worksheet.cell(row=row, column=16, value=report.material_origin)
            worksheet.cell(row=row, column=17, value=report.total_hours)
            worksheet.cell(row=row, column=18, value=report.last_implement_date)
            row += 1  # Increment the row number for the next iteration
        
        for row_number in range(9, 600):
            row_values = [cell.value for cell in worksheet[row_number]]
            # Check if all values in the row are None or empty strings
            if all(value is None or (isinstance(value, str) and value.strip() == '') for value in row_values):
                # Hide the entire row
                worksheet.row_dimensions[row_number].hidden = True
        # Start pasting data from row 100

        # seen_maintenance_scope_names = set()
        # filtered_data_row_100 = []
        # for report in data_row_100:
        #     maintenance_scope_name = report.maintenance_scope_name
        #     if maintenance_scope_name not in seen_maintenance_scope_names:
        #         filtered_data_row_100.append(report)
        #         seen_maintenance_scope_names.add(maintenance_scope_name)
        row = 609  # Adjust the starting row based on the number of rows added to data_row_9
        for report in data_row_100:
            # if row >= 9+9 + len(data_row_9)+14:
            #     worksheet.insert_rows(row)
                # adjust_merged_cells(worksheet, row, 1)
            worksheet.cell(row=row, column=3, value='Hạng mục sửa chữa:  ' + str(report.maintenance_scope_name)
                                                    + '\n' + 'Công việc:  ' + str(report.job_name))
            worksheet.cell(row=row, column=4, value=report.supplier_name)
            worksheet.cell(row=row, column=5, value=report.reason)
            worksheet.cell(row=row, column=6, value=report.material_name)
            worksheet.cell(row=row, column=7, value=report.material_origin)
            worksheet.cell(row=row, column=8, value=report.total_hours)
            worksheet.cell(row=row, column=9, value=report.last_implement_date)
            worksheet.cell(row=row, column=11, value=report.labor_cost)
            worksheet.cell(row=row, column=12, value=report.material_average_quote_price)

            row += 1  # Increment the row number for the next iteration
        for row_number in range(609, 1046):
            row_values = [cell.value for cell in worksheet[row_number]]
            # Check if all values in the row are None or empty strings
            if all(value is None or (isinstance(value, str) and value.strip() == '') for value in row_values):
                # Hide the entire row
                worksheet.row_dimensions[row_number].hidden = True
        # Create a temporary file to store the XLSX data

        # for row_number in range(9, 157) + range(165, 311):
        #     max_len = 0
        #     for cell in worksheet[row_number]:
        #         try:
        #             if len(str(cell.value)) > max_len:
        #                 max_len = len(cell.value)
        #         except:
        #             pass
        #     # Adjust the height (example: you might need to fine-tune the multiplication factor)
        #     worksheet.row_dimensions[row_number].height = max_len * 1,2

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Loop through the worksheet and replace "FALSE" with None
            for row in temp_worksheet.iter_rows():
                for cell in row:
                    if cell.value == False:
                        cell.value = '  '

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, 'rb') as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Báo cáo tuần_{week_number}__{self.date_start}_{self.date_end}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(binary_data),
            'res_model': self._name,
            'res_id': self.id,
            'public': True,
            'res_field': 'custom_export',  # Replace with the appropriate field name
        })

        # Return an action to open the attachment
        return {
            'name': filename,
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}/{filename}',
            'target': 'self',
        }


class JobQuoteReport(models.Model):
    _name = "ship.job.quote.report"
    _description = "Job Quote Report"
    _inherit = ["mail.thread"]
    _check_company_auto = True

    finished_date = fields.Date(string="Finished Date", tracking=True, readonly=True)
    approved_implement_date = fields.Date("Approved implement date", tracking=True)

    maintenance_scope_name=fields.Char(
        tracking=True, readonly=True
    )
    note = fields.Char("Note", tracking=True)
    ship_location = fields.Char("Ship location", tracking=True)
    problem_description = fields.Char("Problem description", tracking=True)
    reason = fields.Char("Reason", tracking=True)
    implement_date = fields.Date("Implement date", tracking=True)
    week_number = fields.Integer(
        string="Week Number", tracking=True
    )
    job_name = fields.Char("Job Name", tracking=True)
    supplier_name = fields.Char(string="Supplier", tracking=True)
    labor_cost = fields.Float("Labor cost", tracking=True)
    material_name = fields.Char("Material Name", tracking=True)
    material_origin = fields.Char("Material Origin", tracking=True)
    material_average_quote_price = fields.Float("material average quote price", tracking=True)
    total_hours = fields.Float(
        "Total usage time in hours", readonly=True, tracking=True
    )
    last_implement_date = fields.Date("Last implement date", tracking=True)


    batch_id = fields.Many2one(
        "ship.job.quote.report.batch", string="Batch", readonly=True
    )

    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    # Add any additional fields and methods that are needed for the report.




