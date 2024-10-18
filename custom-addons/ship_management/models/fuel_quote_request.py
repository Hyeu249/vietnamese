# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from . import CONST
from odoo.exceptions import ValidationError
from datetime import timedelta
from .common_utils import generate_token, format_field_date
import logging
import io
import os
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import tempfile
from odoo.exceptions import UserError
from odoo import _, api, fields, models
from datetime import datetime, timedelta
from openpyxl.styles import Alignment


_logger = logging.getLogger(__name__)


class FuelQuotesRequest(models.Model):
    _name = "ship.fuel.quotes.request"
    _description = "Yêu cầu nhiên liệu"
    _inherit = ["utilities.approval.status"]
    _check_company_auto = True
    _edit_field_permissions_list = {
        "supplier_emails": [],
    }

    order_date = fields.Date("Order date", tracking=True)
    expected_delivery_date = fields.Date("Expected delivery date", tracking=True)
    deadline = fields.Date("Deadline", tracking=True)

    port_id = fields.Many2one("ship.port", string="Port")
    fuel_quote_ids = fields.One2many(
        "ship.fuel.quote", "fuel_quote_request_id", string="Fuel Quote"
    )
    fuel_quote_grease_ids = fields.One2many(
        "ship.fuel.quote.grease", "fuel_quote_request_id", string="Fuel Quote Grease"
    )
    fuel_internal_calculator_id = fields.Many2one(
        "ship.fuel.internal.calculator", string="Fuel internal calculator"
    )
    internal_bunker_request_mt_fo = fields.Float(
        string="Internal Bunker Request F.O (MT)",
        related="fuel_internal_calculator_id.bunker_request_mt_fo",
    )
    internal_bunker_request_mt_do = fields.Float(
        string="Internal Bunker Request D.O (MT)",
        related="fuel_internal_calculator_id.bunker_request_mt_do",
    )
    internal_order_remaining_mt_fo = fields.Float(
        string="Internal Bunker Remaining F.O (MT) At Order",
        compute="_compute_internal_order_remaining_mt_fo",
    )
    internal_order_remaining_mt_do = fields.Float(
        string="Internal Bunker Remaining D.O (MT) At Order",
        compute="_compute_internal_order_remaining_mt_do",
    )

    internal_bunker_remaining_mt_fo = fields.Float(
        string="Internal Bunker Remaining F.O (MT) At Bunkering Port",
        related="fuel_internal_calculator_id.remaining_oil_bunk_port_mt_fo",
    )
    internal_bunker_remaining_mt_do = fields.Float(
        string="Internal Bunker Remaining D.O (MT) At Bunkering Port",
        related="fuel_internal_calculator_id.remaining_oil_bunk_port_mt_do",
    )

    arrival_time = fields.Datetime(string="Arrival Time")
    depature_time = fields.Datetime(string="Departure Time")
    supplier_emails = fields.Char("Supplier emails", readonly=True, tracking=True)

    is_email_sent = fields.Boolean("Is email sent", default=False)
    is_responded = fields.Boolean("Is the supplier responded", default=False)

    # an access_token is a unique string that is automatically generated for each quote.
    access_token = fields.Char("Access Token", default=lambda self: generate_token())

    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    # sequential
    ref = fields.Char(string="Code", default=lambda self: _("New"))
    receipt = fields.Image("Receipt", tracking=True,
                           max_width=CONST.MAX_IMAGE_UPLOAD_WIDTH,
                           max_height=CONST.MAX_IMAGE_UPLOAD_HEIGHT)

    @api.depends(
        "fuel_internal_calculator_id.remaining_oil_serv_tank_fo",
        "fuel_internal_calculator_id.remaining_oil_other_tanks_fo",
    )
    def _compute_internal_order_remaining_mt_fo(self):
        for record in self:
            if record.fuel_internal_calculator_id:
                record.internal_order_remaining_mt_fo = (
                    record.fuel_internal_calculator_id.remaining_oil_serv_tank_fo
                    + record.fuel_internal_calculator_id.remaining_oil_other_tanks_fo
                )
            else:
                record.internal_order_remaining_mt_fo = 0.0

    @api.depends(
        "fuel_internal_calculator_id.remaining_oil_serv_tank_do",
        "fuel_internal_calculator_id.remaining_oil_other_tanks_do",
    )
    def _compute_internal_order_remaining_mt_do(self):
        for record in self:
            if record.fuel_internal_calculator_id:
                record.internal_order_remaining_mt_do = (
                    record.fuel_internal_calculator_id.remaining_oil_serv_tank_do
                    + record.fuel_internal_calculator_id.remaining_oil_other_tanks_do
                )
            else:
                record.internal_order_remaining_mt_do = 0.0

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            model_name = "ship.fuel.quotes.request"
            vals["ref"] = self.env["ir.sequence"].next_by_code(model_name)
        result = super(FuelQuotesRequest, self).create(vals_list)

        for record in result:
            internal_fo = record.internal_bunker_request_mt_fo
            internal_do = record.internal_bunker_request_mt_do
            if internal_fo:
                record._create_fuel_quote("fo", internal_fo)
            if internal_do:
                record._create_fuel_quote("do", internal_do)
        return result

    def _create_fuel_quote(self, name, quantity):
        self.ensure_one()
        self.env["ship.fuel.quote"].create(
            {
                "name": name,
                "quantity": quantity,
                "unit": "MT",
                "fuel_quote_request_id": self.id,
            }
        )

    def name_get(self):
        result = []
        for report in self:
            name = report.ref or _("New")
            result.append((report.id, name))
        return result

    def write(self, vals):
        self.ensure_one()
        result = super(FuelQuotesRequest, self).write(vals)

        if "approval_status" in vals:
            if self.is_at_this_approval_level(CONST.SUPPLIER):
                self._get_supplier_emails()
                for fuel_quote in self.fuel_quote_ids:
                    if not fuel_quote.fuel_supplier_quote_ids:
                        fuel_quote._create_fuel_supplier_quotes()
                for fuel_quote_grease in self.fuel_quote_grease_ids:
                    if not fuel_quote_grease.grease_supplier_quote_ids:
                        fuel_quote_grease._create_fuel_supplier_quotes()

        return result

    def _handle_removed_fuel_quotes(self, removed_ids):
        self.ensure_one()
        if self._are_send_quotes_to_suppliers():
            raise ValidationError(
                "Cannot remove quotes when already sent to suppliers."
            )
        for id in removed_ids:
            fuel_quote_id = self.fuel_quote_ids.browse(id)
            fuel_quote_id.fuel_quotes_request_id = False

    def _handle_added_fuel_quotes(self, added_ids):
        self.ensure_one()
        if self._are_send_quotes_to_suppliers():
            raise ValidationError("Cannot add quotes when already sent to suppliers.")
        for id in added_ids:
            fuel_quote_id = self.fuel_quote_ids.browse(id)
            fuel_quote_id.fuel_quotes_request_id = self.id

    def _get_supplier_emails(self):
        self.ensure_one()
        self.supplier_emails = self.port_id.supplier_id.email

    def get_all_unsent_supplier_quotes(self):
        conditions = [
            ("is_email_sent", "=", False),
        ]
        return self.search(conditions)

    def action_send_emails_to_all_unsent_supplier_quotes(self):
        for record in self.get_all_unsent_supplier_quotes():
            record.action_send_email()

    def action_send_email(self):
        """For sending email to suppliers for each fuel quote."""
        self.ensure_one()
        ##
        for fuel_quote in self.fuel_quote_ids:
            for fuel_supplier_quote_id in fuel_quote.fuel_supplier_quote_ids:
                fuel_supplier_quote_id.is_email_sent = True

        fuel_quotes = (
            self.env["ship.fuel.quote"]
            .sudo()
            .search([("fuel_quote_request_id", "=", self.id)])
        )

        try:
            template = self.env.ref(
                "ship_management.email_template_fuel_quote_request"
            ).id
            base_url = self.env["ir.config_parameter"].sudo().get_param("web.base.url")

            reply_quote_url = (
                "{base_url}/vendor_rfuq/{ref}?access_token={token}".format(
                    base_url=base_url,
                    ref=self.ref,
                    token=self.access_token,
                )
            )
            fuel_quotes = []
            for fuel_quote in self.fuel_quote_ids:
                fuel_quotes.append(
                    {
                        "name": fuel_quote.name,
                        "quantity": fuel_quote.quantity,  # Assuming quantity is stored in the related table
                        "unit": fuel_quote.unit,  # Assuming unit is stored in the related table
                    }
                )

            for fuel_quote_grease in self.fuel_quote_grease_ids:
                fuel_quotes.append(
                    {
                        "name": fuel_quote_grease.name,
                        "quantity": fuel_quote_grease.quantity,  # Assuming quantity is stored in the related table
                        "unit": fuel_quote_grease.unit,  # Assuming unit is stored in the related table
                    }
                )

            context = {
                "supplier_name": self.port_id.supplier_id.name,
                "reply_quote_url": reply_quote_url,
                "fuel_quotes": fuel_quotes,
            }
            email_values = {
                "email_to": self.port_id.supplier_id.email,
            }
            if self.is_at_this_approval_level(CONST.SUPPLIER):
                self.is_email_sent = True
                self.env["mail.template"].browse(template).with_context(
                    context
                ).send_mail(self.id, email_values=email_values, force_send=False)

        except Exception as e:
            self.is_email_sent = False
            logging.error(f"Error sending email for fuel quote request {self.id}: {e}")

    def custom_export_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = "/mnt/extra-addons/report_template/ycnl.xlsx"  # Replace with the actual path
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        worksheet.cell(row=7, column=5, value=self.order_date.strftime("%d/%m/%Y"))

        ###
        row = 12
        index = 0
        for fuel_quote in self.fuel_quote_ids:
            worksheet.insert_rows(row)
            index += 1
            worksheet.cell(row=row, column=1, value=index)
            worksheet.cell(row=row, column=2, value=fuel_quote.full_name)
            worksheet.cell(row=row, column=3, value=fuel_quote.unit)
            worksheet.cell(row=row, column=4, value=fuel_quote.quantity)
            row += 1
        ##
        location_str = (
            f"Địa điểm thực hiện: {self.company_id.name} - {self.port_id.name}"
        )
        worksheet.cell(row=row, column=1, value=location_str)
        ##
        arrival_str = (
            self.arrival_time.strftime("%H.%M GIỜ %d/%m/%Y") + f" {self.port_id.name}"
        )
        departure_str = (
            self.depature_time.strftime("%H.%M GIỜ %d/%m/%Y") + f" {self.port_id.name}"
        )
        worksheet.cell(row=row + 2, column=1, value=arrival_str)
        worksheet.cell(row=row + 3, column=1, value=departure_str)

        worksheet.merge_cells(f"A{row+5}:F{row+5}")

        worksheet.merge_cells(f"C{row+7}:F{row+7}")

        worksheet.merge_cells(f"C{row+8}:F{row+8}")

        align_center = Alignment(horizontal="center", vertical="center")
        worksheet[f"C{row+7}"].alignment = align_center
        worksheet[f"C{row+8}"].alignment = align_center

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Yêu cầu nhiên liệu ngày {self.order_date}.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action
