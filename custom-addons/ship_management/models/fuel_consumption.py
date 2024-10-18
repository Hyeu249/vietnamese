# -*- coding: utf-8 -*-
import io
import os
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import tempfile
from odoo.exceptions import UserError
from odoo import _, api, fields, models
from datetime import datetime, timedelta
import logging
from . import CONST
from ...utilities.models import CONST as UTILITIES_CONST
from odoo.exceptions import ValidationError


_logger = logging.getLogger(__name__)


class ShipFuelConsumptionReportBatch(models.Model):
    _name = "ship.fuel.consumption.report.batch"
    _description = "Ship Fuel Consumption Report Batch"
    _check_company_auto = True

    name = fields.Char("Name")
    date_start = fields.Date("Date From", required=True)
    date_end = fields.Date("Date To", required=True)
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )
    line_ids = fields.One2many(
        "ship.fuel.consumption.report", "batch_id", string="Report Lines"
    )

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals["name"]:
                model_code = "ship.fuel.consumption.report.batch"
                vals["name"] = self.env["ir.sequence"].next_by_code(model_code)
        return super(ShipFuelConsumptionReportBatch, self).create(vals_list)

    def delete_temporary_attachments(self):
        # Find all attachments marked as temporary
        temporary_attachments = self.env["ir.attachment"].search(
            [("res_field", "=", "custom_export")]
        )
        # Delete them
        for attachment in temporary_attachments:
            attachment.unlink()

    def generate_report(self):
        self.ensure_one()
        self.env["ship.fuel.consumption.report"].search(
            [("batch_id", "=", self.id)]
        ).unlink()
        current_company_id = self.env.company.id
        query = """
                
                
            WITH filtered_ship_fuel AS (
                        SELECT 
                            *,
                            time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh' as local_time
                        FROM 
                            ship_fuel
                        WHERE 
                            (time BETWEEN %s AND %s)  AND
                            (company_id = %s)
                    )


            SELECT 
                    voy,
                    local_time::date as local_date,
                    TO_CHAR(local_time, 'HH24:MI') as local_time_hour_minute,
                    voy_status,
                    voy_position,
                    0 as time_diff_hours,
                    0 as fo_diff,
                    0 as do_diff,
                    0 as cyl_diff,
                    0 as lome_diff,
                    0 as loge_diff,
                    me_rpm,
                    load,
                    tc_rpm,
                    pa,
                    spd,
                    distance_run,
                    weather
                FROM 
                    (SELECT 
                        voy,
                        local_time,
                        voy_status,
                        voy_position,
                        current_fo,
                        current_do,
                        current_cyl,
                        current_lome,
                        current_loge,
                        me_rpm,
                        load,
                        tc_rpm,
                        pa,
                        spd,
                        distance_run,
                        weather,
                        ROW_NUMBER() OVER (PARTITION BY voy ORDER BY local_time) as rn
                    FROM 
                        filtered_ship_fuel) as sub
                WHERE 
                    rn = 1

                UNION

                SELECT 
                    voy,
                    LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time)::date AS local_date,
                    TO_CHAR(LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time), 'HH24:MI') AS local_time_hour_minute,
                    LEAD(voy_status) OVER (PARTITION BY voy ORDER BY local_time) AS voy_status,
                    LEAD(voy_position) OVER (PARTITION BY voy ORDER BY local_time) AS voy_position,
                    EXTRACT(EPOCH FROM (LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time) - local_time)) / 3600.0 AS time_diff_hours,
                    current_fo - LEAD(current_fo) OVER (PARTITION BY voy ORDER BY local_time) AS fo_diff,
                    current_do - LEAD(current_do) OVER (PARTITION BY voy ORDER BY local_time) AS do_diff,
                    current_cyl - LEAD(current_cyl) OVER (PARTITION BY voy ORDER BY local_time) AS cyl_diff,
                    current_lome - LEAD(current_lome) OVER (PARTITION BY voy ORDER BY local_time) AS lome_diff,
                    current_loge - LEAD(current_loge) OVER (PARTITION BY voy ORDER BY local_time) AS loge_diff,
                    LEAD(me_rpm) OVER (PARTITION BY voy ORDER BY local_time) AS me_rpm,
                    LEAD(load) OVER (PARTITION BY voy ORDER BY local_time) AS load,
                    LEAD(tc_rpm) OVER (PARTITION BY voy ORDER BY local_time) AS tc_rpm,
                    LEAD(pa) OVER (PARTITION BY voy ORDER BY local_time) AS pa,
                    LEAD(spd) OVER (PARTITION BY voy ORDER BY local_time) AS spd,
                    LEAD(distance_run) OVER (PARTITION BY voy ORDER BY local_time) AS distance_run,
                    LEAD(weather) OVER (PARTITION BY voy ORDER BY local_time) AS weather
                FROM 
                    filtered_ship_fuel
                ORDER BY 
                    local_date,
                    local_time_hour_minute,
                    voy_status
                            
                ;
			
	
                """

        self.env.cr.execute(query, (self.date_start, self.date_end,current_company_id))
        results = self.env.cr.fetchall()

        data = []
        for result in results:
            if result[3] is not None:
                data.append(
                    {
                        "voy": result[0],
                        "local_date": result[
                            1
                        ],  # Changed from local_time to local_date
                        "local_time_hour_minute": result[2],  # Added new field
                        "voy_status": result[3],
                        "voy_position": result[4],
                        "time_diff_hours": result[5],
                        "fo_diff": result[6],
                        "do_diff": result[7],
                        "cyl_diff": result[8],
                        "lome_diff": result[9],
                        "loge_diff": result[10],
                        "me_rpm": result[11],
                        "load": result[12],
                        "tc_rpm": result[13],
                        "pa": result[14],
                        "spd": result[15],
                        "distance_run": result[16],
                        "weather": result[17],  # Adjusted index due to new fields
                        "batch_id": self.id,
                    }
                )

        self.env["ship.fuel.consumption.report"].create(data)

    def custom_export_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = (
            "/mnt/extra-addons/report_template/bcd.xlsx"  # Replace with the actual path
        )
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        # week_number = self.date_start.isocalendar()[1]

        # worksheet['A1']= f"BÁO CÁO GIAO BAN TUẦN {week_number}"
        # worksheet['D4']= f"{week_number}"
        # worksheet['D161']= f"{week_number+1}"
        row = 12
        for report in self.line_ids:
            if report.voy_status is not None:
                worksheet.cell(row=row, column=13, value=report.local_date)
                worksheet.cell(row=row, column=14, value=report.voy)
                worksheet.cell(row=row, column=15, value=report.voy_position)
                worksheet.cell(row=row, column=16, value=report.local_time_hour_minute)
                if "FE" in report.voy_status:
                    worksheet.cell(row=row, column=17, value="FE")
                elif "SB" in report.voy_status:
                    worksheet.cell(row=row, column=17, value="SB")
                elif "RU" in report.voy_status:
                    worksheet.cell(row=row, column=17, value="RU")
                else:
                    worksheet.cell(row=row, column=17, value=report.voy_status)

                if report.voy_status == "DEP_SB":
                    # Populate specific columns for 'SB'
                    worksheet.cell(row=row, column=18, value=report.time_diff_hours)
                    worksheet.cell(
                        row=row,
                        column=19,
                        value=(
                            float(report.fo_diff)
                            if report.fo_diff == 0
                            else report.fo_diff
                        ),
                    )
                    worksheet.cell(
                        row=row,
                        column=20,
                        value=(
                            float(report.do_diff)
                            if report.do_diff == 0
                            else report.do_diff
                        ),
                    )
                elif report.voy_status == "ANCH_ARV_SB":
                    # Populate specific columns for 'SB'
                    worksheet.cell(row=row, column=24, value=report.time_diff_hours)
                    worksheet.cell(
                        row=row,
                        column=25,
                        value=(
                            float(report.fo_diff)
                            if report.fo_diff == 0
                            else report.fo_diff
                        ),
                    )
                    worksheet.cell(
                        row=row,
                        column=26,
                        value=(
                            float(report.do_diff)
                            if report.do_diff == 0
                            else report.do_diff
                        ),
                    )
                elif report.voy_status == "ANCH_FE":
                    worksheet.cell(row=row, column=21, value=report.time_diff_hours)
                    worksheet.cell(
                        row=row,
                        column=22,
                        value=(
                            float(report.fo_diff)
                            if report.fo_diff == 0
                            else report.fo_diff
                        ),
                    )
                    worksheet.cell(
                        row=row,
                        column=23,
                        value=(
                            float(report.do_diff)
                            if report.do_diff == 0
                            else report.do_diff
                        ),
                    )
                elif (
                    report.voy_status == "ARV_SB"
                    and last_report.voy_status == "ANCH_FE"
                ):
                    worksheet.cell(row=row, column=18, value=report.time_diff_hours)
                    worksheet.cell(
                        row=row,
                        column=19,
                        value=(
                            float(report.fo_diff)
                            if report.fo_diff == 0
                            else report.fo_diff
                        ),
                    )
                    worksheet.cell(
                        row=row,
                        column=20,
                        value=(
                            float(report.do_diff)
                            if report.do_diff == 0
                            else report.do_diff
                        ),
                    )
                elif "RU" in report.voy_status or report.voy_status == "ARV_FE":
                    # Populate specific columns for 'RU'
                    # Replace with actual column values for 'RU'
                    worksheet.cell(row=row, column=21, value=report.time_diff_hours)
                    worksheet.cell(
                        row=row,
                        column=22,
                        value=(
                            float(report.fo_diff)
                            if report.fo_diff == 0
                            else report.fo_diff
                        ),
                    )
                    worksheet.cell(
                        row=row,
                        column=23,
                        value=(
                            float(report.do_diff)
                            if report.do_diff == 0
                            else report.do_diff
                        ),
                    )
                elif report.voy_status == "NOON" or report.voy_status == "ARV_SB":
                    # Populate specific columns for 'NOON'
                    # Replace with actual column values for 'NOON'
                    worksheet.cell(row=row, column=24, value=report.time_diff_hours)
                    worksheet.cell(
                        row=row,
                        column=25,
                        value=(
                            float(report.fo_diff)
                            if report.fo_diff == 0
                            else report.fo_diff
                        ),
                    )
                    worksheet.cell(
                        row=row,
                        column=26,
                        value=(
                            float(report.do_diff)
                            if report.do_diff == 0
                            else report.do_diff
                        ),
                    )

                worksheet.cell(row=row, column=27, value=report.me_rpm)
                worksheet.cell(row=row, column=28, value=report.load)
                worksheet.cell(row=row, column=29, value=report.tc_rpm)
                worksheet.cell(row=row, column=30, value=report.pa)
                worksheet.cell(row=row, column=31, value=report.spd)
                worksheet.cell(row=row, column=32, value=report.distance_run)
                worksheet.cell(row=row, column=33, value=report.weather)

                row += 1
                last_report = report

        current_company_id = self.env.company.id

        ###########
        query = """
                
                
            WITH filtered_ship_fuel AS (
                        SELECT 
                            *,
                            time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh' as local_time
                        FROM 
                            ship_fuel
                        WHERE 
                            (time BETWEEN %s AND %s)  AND
                            (company_id = %s)
                    ),


                voy_data as (SELECT 
                        voy,
                        local_time,
                        local_time::date as local_date,
                        TO_CHAR(local_time, 'HH24:MI') as local_time_hour_minute,
                        voy_status,
                        voy_position,
                        0 as time_diff_hours,
                        0 as fo_diff,
                        0 as do_diff,
                        0 as cyl_diff,
                        0 as lome_diff,
                        0 as loge_diff,
                        me_rpm,
                        load,
                        tc_rpm,
                        pa,
                        spd,
                        distance_run,
                        weather
                    FROM 
                        (SELECT 
                            voy,
                            local_time,
                            voy_status,
                            voy_position,
                            current_fo,
                            current_do,
                            current_cyl,
                            current_lome,
                            current_loge,
                            me_rpm,
                            load,
                            tc_rpm,
                            pa,
                            spd,
                            distance_run,
                            weather,
                            ROW_NUMBER() OVER (PARTITION BY voy ORDER BY local_time) as rn

                        FROM 
                            filtered_ship_fuel) as sub
                    WHERE 
                        rn = 1

                    UNION

                    SELECT 
                        voy,
                        local_time,
                        LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time)::date AS local_date,
                        TO_CHAR(LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time), 'HH24:MI') AS local_time_hour_minute,
                        LEAD(voy_status) OVER (PARTITION BY voy ORDER BY local_time) AS voy_status,
                        LEAD(voy_position) OVER (PARTITION BY voy ORDER BY local_time) AS voy_position,
                        EXTRACT(EPOCH FROM (LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time) - local_time)) / 3600.0 AS time_diff_hours,
                        current_fo - LEAD(current_fo) OVER (PARTITION BY voy ORDER BY local_time) AS fo_diff,
                        current_do - LEAD(current_do) OVER (PARTITION BY voy ORDER BY local_time) AS do_diff,
                        current_cyl - LEAD(current_cyl) OVER (PARTITION BY voy ORDER BY local_time) AS cyl_diff,
                        current_lome - LEAD(current_lome) OVER (PARTITION BY voy ORDER BY local_time) AS lome_diff,
                        current_loge - LEAD(current_loge) OVER (PARTITION BY voy ORDER BY local_time) AS loge_diff,
                        LEAD(me_rpm) OVER (PARTITION BY voy ORDER BY local_time) AS me_rpm,
                        LEAD(load) OVER (PARTITION BY voy ORDER BY local_time) AS load,
                        LEAD(tc_rpm) OVER (PARTITION BY voy ORDER BY local_time) AS tc_rpm,
                        LEAD(pa) OVER (PARTITION BY voy ORDER BY local_time) AS pa,
                        LEAD(spd) OVER (PARTITION BY voy ORDER BY local_time) AS spd,
                        LEAD(distance_run) OVER (PARTITION BY voy ORDER BY local_time) AS distance_run,
                        LEAD(weather) OVER (PARTITION BY voy ORDER BY local_time) AS weather
                    FROM 
                        filtered_ship_fuel
                    ORDER BY 
                        local_date,
                        local_time_hour_minute,
                        voy_status),
                        
                processed_voy_data as (select *,        LAG(voy_status) OVER (PARTITION BY voy ORDER BY local_time) AS prev_voy_status
            from voy_data)
                        
                        SELECT 
                            LEFT(voy, 4) as voy_group,
                            SUM(CASE WHEN voy_status IN ('NOON', 'ARV_SB', 'ANCH_ARV_SB') and prev_voy_status != 'ANCH_FE' THEN time_diff_hours ELSE 0 END) as sea_time_diff,
                            SUM(CASE WHEN voy_status NOT IN ('NOON', 'ARV_SB', 'ANCH_ARV_SB') THEN time_diff_hours 
                                    WHEN voy_status= 'ARV_SB' and prev_voy_status = 'ANCH_FE' THEN time_diff_hours ELSE 0 END) as anchor_time_diff,
                            SUM(distance_run) as total_distance,
                            AVG(CASE WHEN spd > 0 THEN spd ELSE NULL END) as avg_speed,
                            SUM(fo_diff) as total_fo_diff,
                            SUM(do_diff) as total_do_diff,
                            SUM(cyl_diff) as total_cyl_diff,
                            SUM(lome_diff) as total_lome_diff,
                            SUM(loge_diff) as total_loge_diff
                        FROM 
                            processed_voy_data
                        GROUP BY 
                            LEFT(voy, 4)
                        ORDER BY 
                            LEFT(voy, 4);

	
                """
        self.env.cr.execute(query, (self.date_start, self.date_end,current_company_id))
        sql_results = self.env.cr.fetchall()

        row = 12
        for result in sql_results:
            worksheet.cell(row=row, column=2, value=result[0])  # value1 in column 1
            worksheet.cell(row=row, column=3, value=result[1])
            worksheet.cell(row=row, column=4, value=result[2])
            worksheet.cell(row=row, column=5, value=result[3])
            worksheet.cell(row=row, column=6, value=result[4])
            worksheet.cell(row=row, column=7, value=result[5])
            worksheet.cell(row=row, column=8, value=result[6])
            worksheet.cell(row=row, column=9, value=result[7])
            worksheet.cell(row=row, column=10, value=result[8])
            worksheet.cell(row=row, column=11, value=result[9])

            row += 1  # Move to the next row for the next record

        #################
        query = """
                
                
            WITH filtered_ship_fuel AS (
                        SELECT 
                            *,
                            time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Ho_Chi_Minh' as local_time
                        FROM 
                            ship_fuel
                        WHERE 
                            (time BETWEEN %s AND %s)  AND
                            (company_id = %s)
                    ),
                voy_data as (SELECT 
                                    voy,
                                    local_time::date as local_date,
                                    TO_CHAR(local_time, 'HH24:MI') as local_time_hour_minute,
                                    voy_status,
                                    voy_position,
                                    0 as time_diff_hours,
                                    0 as fo_diff,
                                    0 as do_diff,
                                    0 as cyl_diff,
                                    0 as lome_diff,
                                    0 as loge_diff,
                                    me_rpm,
                                    load,
                                    tc_rpm,
                                    pa,
                                    spd,
                                    distance_run,
                                    weather
                                FROM 
                                    (SELECT 
                                        voy,
                                        local_time,
                                        voy_status,
                                        voy_position,
                                        current_fo,
                                        current_do,
                                        current_cyl,
                                        current_lome,
                                        current_loge,
                                        me_rpm,
                                        load,
                                        tc_rpm,
                                        pa,
                                        spd,
                                        distance_run,
                                        weather,
                                        ROW_NUMBER() OVER (PARTITION BY voy ORDER BY local_time) as rn
                                    FROM 
                                        filtered_ship_fuel) as sub
                                WHERE 
                                    rn = 1

                                UNION

                                SELECT 
                                    voy,
                                    LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time)::date AS local_date,
                                    TO_CHAR(LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time), 'HH24:MI') AS local_time_hour_minute,
                                    LEAD(voy_status) OVER (PARTITION BY voy ORDER BY local_time) AS voy_status,
                                    LEAD(voy_position) OVER (PARTITION BY voy ORDER BY local_time) AS voy_position,
                                    EXTRACT(EPOCH FROM (LEAD(local_time) OVER (PARTITION BY voy ORDER BY local_time) - local_time)) / 3600.0 AS time_diff_hours,
                                    current_fo - LEAD(current_fo) OVER (PARTITION BY voy ORDER BY local_time) AS fo_diff,
                                    current_do - LEAD(current_do) OVER (PARTITION BY voy ORDER BY local_time) AS do_diff,
                                    current_cyl - LEAD(current_cyl) OVER (PARTITION BY voy ORDER BY local_time) AS cyl_diff,
                                    current_lome - LEAD(current_lome) OVER (PARTITION BY voy ORDER BY local_time) AS lome_diff,
                                    current_loge - LEAD(current_loge) OVER (PARTITION BY voy ORDER BY local_time) AS loge_diff,
                                    LEAD(me_rpm) OVER (PARTITION BY voy ORDER BY local_time) AS me_rpm,
                                    LEAD(load) OVER (PARTITION BY voy ORDER BY local_time) AS load,
                                    LEAD(tc_rpm) OVER (PARTITION BY voy ORDER BY local_time) AS tc_rpm,
                                    LEAD(pa) OVER (PARTITION BY voy ORDER BY local_time) AS pa,
                                    LEAD(spd) OVER (PARTITION BY voy ORDER BY local_time) AS spd,
                                    LEAD(distance_run) OVER (PARTITION BY voy ORDER BY local_time) AS distance_run,
                                    LEAD(weather) OVER (PARTITION BY voy ORDER BY local_time) AS weather
                                FROM 
                                    filtered_ship_fuel
                                ORDER BY 
                                    local_date,
                                    local_time_hour_minute,
                                    voy_status),

                total_voy as (SELECT 
                    LEFT(voy, 4) as voy_group,
                    SUM(CASE WHEN voy_status IN ('NOON', 'ARV_SB', 'ANCH_ARV_SB') THEN time_diff_hours ELSE 0 END) as sea_time_diff,
                    SUM(CASE WHEN voy_status NOT IN ('NOON', 'ARV_SB', 'ANCH_ARV_SB') THEN time_diff_hours ELSE 0 END) as anchor_time_diff,
                    SUM(distance_run) as total_distance,
                    AVG(CASE WHEN spd > 0 THEN spd ELSE NULL END) as avg_speed,
                    SUM(fo_diff) as total_fo_diff,
                    SUM(do_diff) as total_do_diff,
                    SUM(cyl_diff) as total_cyl_diff,
                    SUM(lome_diff) as total_lome_diff,
                    SUM(loge_diff) as total_loge_diff
                FROM 
                    voy_data
                GROUP BY 
                    LEFT(voy, 4)
                ORDER BY 
                    LEFT(voy, 4))
                    
                select 
                    count(voy_group),
                    sum(sea_time_diff),
                    sum(anchor_time_diff),
                    SUM(total_distance), 
                    AVG(avg_speed) ,
                    SUM(total_fo_diff) ,
                    SUM(total_do_diff) ,
                    SUM(total_cyl_diff) ,
                    SUM(total_lome_diff) ,
                    SUM(total_loge_diff) 
                    from total_voy
                    
                        ;
                                            
                
			
	
                """
        self.env.cr.execute(query, (self.date_start, self.date_end,current_company_id))
        sql_results = self.env.cr.fetchall()

        row = 5
        for result in sql_results:
            worksheet.cell(row=row, column=1, value=result[0])  # value1 in column 1
            worksheet.cell(row=row, column=2, value=result[1])
            worksheet.cell(row=row, column=3, value=result[2])
            worksheet.cell(row=row, column=4, value=result[3])
            worksheet.cell(row=row, column=5, value=result[4])
            worksheet.cell(row=row, column=6, value=result[5])
            worksheet.cell(row=row, column=7, value=result[6])
            worksheet.cell(row=row, column=8, value=result[7])
            worksheet.cell(row=row, column=9, value=result[8])
            worksheet.cell(row=row, column=10, value=result[9])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            for row in temp_worksheet.iter_rows():
                for cell in row:
                    if cell.value == False:
                        # Get column index (1-based)
                        column_index = cell.column

                        # Check if column index is not between 18 and 26
                        if not (18 <= column_index <= 26) and not (
                            2 <= column_index <= 11
                        ):
                            cell.value = None

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Báo cáo từ__{self.date_start}_{self.date_end}.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action


class ShipFuelConsumptionReport(models.Model):
    _name = "ship.fuel.consumption.report"
    _description = "Ship Fuel Consumption Report"
    _check_company_auto = True

    voy = fields.Char("Voyage", tracking=True, readonly=True)
    local_date = fields.Date("Local Date", tracking=True, readonly=True)
    local_time_hour_minute = fields.Char(
        "Local time hour minutes", tracking=True, readonly=True
    )
    voy_status = fields.Char("Voyage Status", tracking=True, readonly=True)
    voy_position = fields.Char("Voyage Position", tracking=True, readonly=True)
    time_diff_hours = fields.Float("Time Difference (Hours)")
    fo_diff = fields.Float(
        "Fuel Oil Difference",
    )
    do_diff = fields.Float("Diesel Oil Difference")
    cyl_diff = fields.Float("Cylinder Oil Difference", tracking=True, readonly=True)
    lome_diff = fields.Float("Lome Difference", tracking=True, readonly=True)
    loge_diff = fields.Float("Loge Difference", tracking=True, readonly=True)
    me_rpm = fields.Float("ME RPM", tracking=True, readonly=True)
    load = fields.Char("Load", tracking=True, readonly=True)
    tc_rpm = fields.Float("TC RPM", tracking=True, readonly=True)
    pa = fields.Char("PA", tracking=True, readonly=True)
    spd = fields.Float("Speed", tracking=True, readonly=True)
    distance_run = fields.Float("Distance Run", tracking=True, readonly=True)
    weather = fields.Char("Weather", tracking=True, readonly=True)
    batch_id = fields.Many2one(
        "ship.fuel.consumption.report.batch", string="Batch", readonly=True
    )
    notes = fields.Text("Notes")
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )


class ShipFuelConsumption(models.Model):
    _name = "ship.fuel"
    _description = "Ship Fuel"
    _inherit = ["utilities.notification"]

    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )
    voy = fields.Char("Voyage", required=True)
    status = fields.Selection(
        [("NOON_RPT", "NOON RPT"), ("ARV_RPT", "ARV RPT"), ("DEPT_RPT", "DEPT RPT")],
        string="Status",
        required=True,
        default="NOON_RPT",
    )
    voy_status = fields.Selection(
        [
            ("DEP_FE", "DEP FE"),
            ("DEP_SB", "DEP SB"),
            ("DEP_RU", "DEP RU"),
            ("NOON", "NOON"),
            ("ANCH_FE", "ANCH FE"),
            ("ANCH_ARV_SB", "ANCH ARV SB"),
            ("ANCH_DEP_SB", "ANCH DEP SB"),
            ("ANCH_RU", "ANCH RU"),
            ("ARV_SB", "ARV SB"),
            ("ARV_FE", "ARV FE"),
        ],
        string="Voy Status",
        required=True,
        default="NOON",
    )
    voy_position = fields.Selection(
        [
            ("VICT", "VICT- HỒ CHÍ MINH"),
            ("PTSC", "PTSC ĐÌNH VŨ - HẢI PHÒNG"),
        ],
        string="Position",
    )
    time = fields.Datetime(string="Timestamp")
    me_rpm = fields.Float("ME RPM")
    load = fields.Char("Load")
    tc_rpm = fields.Float("TC RPM")
    pa = fields.Char("PA")
    spd = fields.Float("speed")
    distance_run = fields.Float("distance run")
    weather = fields.Char("Weather")

    # Fields for Current Fuel (used across all status types)
    current_fo = fields.Float("Current FO")
    current_do = fields.Float("Current DO")
    current_cyl = fields.Float("Current CYL")
    current_lome = fields.Float("Current LOME")
    current_loge = fields.Float("Current LOGE")
    re_fo = fields.Float("Received FO", default=0)
    re_do = fields.Float("Received DO", default=0)
    re_cyl = fields.Float("Received CYL", default=0)
    re_lome = fields.Float("Received LOME", default=0)
    re_loge = fields.Float("Received LOGE", default=0)

    def _get_default_value_model(self):
        model_name = "utilities.default.value"
        default_value_model = self.env[model_name].search([])

        return default_value_model

    def _get_current_fo(self):
        default_value_model = self._get_default_value_model()
        variable_name = UTILITIES_CONST.INTEGER_SHIP_FUEL_CONSUMPTION_CURRENT_FO
        return default_value_model._get_default_value_by_variable_name(variable_name)

    def _get_current_do(self):
        default_value_model = self._get_default_value_model()
        variable_name = UTILITIES_CONST.INTEGER_SHIP_FUEL_CONSUMPTION_CURRENT_DO
        return default_value_model._get_default_value_by_variable_name(variable_name)

    def check_fuel_threshold_and_notify(self):
        default_fo = self._get_current_fo()
        default_do = self._get_current_do()
        # Define your fuel thresholds
        thresholds = {
            "current_fo": default_fo,  # Threshold for Fuel Oil (FO)
            "current_do": default_do,  # Threshold for Diesel Oil (DO)
        }
        latest_record = self.search([], order="time desc", limit=1)
        if not latest_record:
            raise UserError(_("No fuel records found."))

        # Check if any fuel is below the threshold and notify

        for value, threshold in thresholds.items():
            current_value = getattr(latest_record, value, 0)
            # Use getattr to dynamically get the field value
            if current_value < threshold:
                self.send_notification_threshold(value, current_value)

    def send_notification_threshold(self, fuel_type="current_fo", current_value=100):
        channel = self.env["mail.channel"].search(
            [("name", "=", "Fuel Notifications")], limit=1
        )
        if channel:
            self._console_log(f"Channel found: {channel.name}")
            try:
                single_channel = channel[0]
                subject = "Fuel Threshold Alert"
                message = f"Nhiên liệu {fuel_type} cần được đặt thêm Giá trị hiện tại: {current_value}"
                # Post a message in the channel
                single_channel.message_post(
                    body=message,
                    subject=subject,
                    message_type="comment",
                    subtype_xmlid="mail.mt_comment",
                )
                self._console_log(f"Sent")
            except:
                self._console_log(f"Fail")
        else:
            self._console_log("Channel not found")
            channel = self.env["mail.channel"].create(
                {
                    "name": "Fuel Notifications",
                    "description": "Notifications for Monthly Fuel Orders",
                    "channel_type": "channel",
                }
            )
            subject = "Fuel Threshold Alert"
            message = f"Chuyến Nội Địa: Nhiên liệu {fuel_type} cần được đặt thêm Giá trị hiện tại: {current_value}"
            # Post a message in the channel
            channel.message_post(
                body=message,
                subject=subject,
                message_type="comment",
                subtype_xmlid="mail.mt_comment",
            )

    def _get_default_day_of_the_month(self):
        default_value_model = self._get_default_value_model()
        variable_name = UTILITIES_CONST.INTEGER_SHIP_FUEL_DAY_OF_THE_MONTH
        return default_value_model._get_default_value_by_variable_name(variable_name)

    def monthly_fuel_order_check(self):
        today = fields.Date.today()
        if (
            today.day == self._get_default_day_of_the_month()
        ):  # Check if today is the 5th day of the month
            self.send_monthly_order_notification()

    def send_monthly_order_notification(self):
        channel = self.env["mail.channel"].search(
            [("name", "=", "Fuel Notifications")], limit=1
        )
        if not channel:
            channel = self.env["mail.channel"].create(
                {
                    "name": "Fuel Notifications",
                    "description": "Notifications for Monthly Fuel Orders",
                    "channel_type": "channel",
                }
            )

        subject = "Monthly Fuel Order Reminder"
        message = "Dầu Bôi Trơn cần được đặt thêm"

        # Post a message in the channel
        channel.message_post(
            body=message,
            subject=subject,
            message_type="comment",
            subtype_xmlid="mail.mt_comment",
        )
