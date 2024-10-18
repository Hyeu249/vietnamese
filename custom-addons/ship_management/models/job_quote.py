# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from . import CONST
from ...utilities.models import CONST as CONST_UTILITIES
import logging
from odoo.exceptions import ValidationError
from datetime import timedelta
import base64
from io import BytesIO
from openpyxl import load_workbook
import logging
from . import listen

_logger = logging.getLogger(__name__)


class JobQuoteTemplate(models.Model):
    _name = "ship.job.quote.template"
    _description = "Báo giá công việc"
    _inherit = ["utilities.approval.status", "ship.date"]
    _edit_field_permissions_list = {
        "deadline": [],
        "note": [],
        "front_image": [],
        "back_image": [],
        "acceptance_quality_level": [],
        "expected_delivery_date": [],
        "is_for_crew": [],
        "implement_date": [],
        "finished_at": [],
        "is_selected_supplier_informed": [],
        "average_quote_price": [],
        "approved_implement_date": [],
        "view_type": [],
        "week_number": [],
        "labor_cost": [],
        "request_state": [],
        "job_state": [],
        "progress_by_date": [],
        "priority": [],
        "ribbon_color": [],
        "maintenance_scope_name": [],
        "ref": [],
        "company_id": [],
        "maintenance_scope_report_id": [],
        "job_id": [],
        # "job_supplier_quote_id": [],
        "job_supplier_quote_ids": [],
        "material_assignment_ids": [],
        "area_of_paint_job_ids": [],
        "inspection_date": [],
    }

    deadline = fields.Date("Deadline", tracking=True)
    note = fields.Char("Note", tracking=True)
    front_image = fields.Image("Front image", tracking=True,
                               max_width=CONST.MAX_IMAGE_UPLOAD_WIDTH,
                               max_height=CONST.MAX_IMAGE_UPLOAD_HEIGHT)
    back_image = fields.Image("Back image", tracking=True,
                              max_width=CONST.MAX_IMAGE_UPLOAD_WIDTH,
                              max_height=CONST.MAX_IMAGE_UPLOAD_HEIGHT)
    acceptance_quality_level = fields.Char("Acceptance quality level", tracking=True)
    expected_delivery_date = fields.Date("Expected delivery date", tracking=True)
    is_for_crew = fields.Boolean("Is for crew", default=True, tracking=True)
    implement_date = fields.Date("Implement date", required=True, tracking=True)
    finished_at = fields.Date(
        "Finished at",
        readonly=lambda self: not self.user.has_group("utilities.group_ship_admin"),
        tracking=True,
    )
    is_selected_supplier_informed = fields.Boolean(
        "Is selected supplier informed", readonly=True
    )
    approved_implement_date = fields.Date("Approved implement date", tracking=True)
    request_state = fields.Selection(
        CONST.JOB_QUOTE_REQUEST_STATES,
        default=CONST.PREPARE,
        string="Request state",
        readonly=True,
        tracking=True,
    )
    job_state = fields.Selection(
        CONST.JOB_STATES,
        string="Job state",
        default=CONST.UNAPPROVED,
        group_expand="_expand_groups",
        tracking=True,
    )

    @api.model
    def _expand_groups(self, states, domain, order):
        return [
            CONST.UNAPPROVED,
            CONST.TODO,
            CONST.WORKING,
            CONST.COMPLETED,
            CONST.CONFIRMED,
        ]

    # compute fields
    progress_by_date = fields.Selection(
        CONST.SCHEDULE_STATES,
        string="Job sub state",
        compute="_set_progress_by_date",
        tracking=True,
    )

    priority = fields.Selection(
        CONST.PRIORITY_SELECTION,
        string="Priority",
        compute="_get_priority_value",
        tracking=True,
    )

    ribbon_color = fields.Integer(
        string="Ribbon color",
        compute="_set_ribbon_color_value",
        default=CONST.YELLOW_CODE,
    )
    view_type = fields.Boolean("View type", tracking=True)
    week_number = fields.Integer(string="Week Number")
    average_quote_price = fields.Float(
        "Average quote price", compute="_calc_average_quote_price", tracking=True
    )

    # relations
    maintenance_scope_report_id = None
    job_id = None
    job_supplier_quote_id = fields.Many2one(
        "ship.job.supplier.quote", string="selected job supplier", tracking=True
    )
    job_supplier_quote_ids = []
    material_assignment_ids = []
    area_of_paint_job_ids = []

    # listen
    @listen.job_supplier_quote.on_write("labor_cost")
    def listen_labor_cost_in_job_quote_template(self):
        for record in self:
            record.propose_when_all_suppliers_have_priced()

    # computes
    @api.depends("implement_date")
    def _get_priority_value(self):
        for record in self:
            if record.implement_date:
                if record.implement_date <= fields.Date.today():
                    record.priority = CONST.HIGH
                else:
                    record.priority = CONST.LOW
            record.priority = CONST.LOW

    @api.depends("implement_date", "job_state")
    def _set_ribbon_color_value(self):
        for record in self:
            implement_date = record.implement_date
            job_state = record.job_state
            today = fields.Date.today()
            if implement_date:
                if implement_date > today:
                    record.ribbon_color = CONST.GREEN_CODE
                elif implement_date == today:
                    record.ribbon_color = CONST.YELLOW_CODE
                elif implement_date < today:
                    if job_state in [CONST.WORKING, CONST.COMPLETED, CONST.CONFIRMED]:
                        record.ribbon_color = None
                    else:
                        record.ribbon_color = CONST.RED_CODE
            else:
                record.ribbon_color = None

    @api.depends("implement_date", "job_state")
    def _set_progress_by_date(self):
        for record in self:
            implement_date = record.implement_date
            job_state = record.job_state
            today = fields.Date.today()
            if implement_date:
                if implement_date > today:
                    record.progress_by_date = CONST.UNDUE
                elif implement_date == today:
                    record.progress_by_date = CONST.TODAY
                elif implement_date < today:
                    if job_state in [CONST.WORKING, CONST.COMPLETED, CONST.CONFIRMED]:
                        record.progress_by_date = CONST.BLANK
                    else:
                        record.progress_by_date = CONST.OVERDUE
            else:
                record.progress_by_date = None

    @api.depends("job_supplier_quote_id", "implement_date")
    def _calc_average_quote_price(self):
        for record in self:
            if record.implement_date and isinstance(record.id, int):
                job_quote_ids = record.get_the_last_3_times_of_this_quote_for_the_job()
                job_supplier_quote_ids = job_quote_ids.mapped("job_supplier_quote_id")
                labor_costs = job_supplier_quote_ids.mapped("labor_cost")

                total = sum(labor_costs)
                count = len(labor_costs)
                if total and count:
                    average_price = total / count
                    record.average_quote_price = average_price
                else:
                    record.average_quote_price = 0
            else:
                record.average_quote_price = 0

    @api.constrains("maintenance_scope_report_id", "job_id")
    def _check_if_maintenance_scope_has_this_job(self):
        message = "Công việc của báo giá, và báo cáo sửa chữa không liên quan đến nhau!"
        for record in self:
            scope_id = record.maintenance_scope_report_id.maintenance_scope_id
            job_id = record.job_id

            if scope_id and job_id:
                if job_id not in scope_id.job_ids:
                    raise ValidationError(message)

    @api.model_create_multi
    def create(self, vals_list):
        result = super(JobQuoteTemplate, self).create(vals_list)

        for record in result:
            record.set_approval_status_selection_based_on_is_for_crew()
            record.approve_quote_if_for_crew()
            record.set_job_state_is_todo_if_job_quote_is_approved()
            record.set_approved_implement_date_if_approve_job_quote()
            record._set_deadline_based_on_expected_delivery_date()
            record._compute_week_number()

        listen.job_quote.call_on_create(result)
        return result

    def write(self, vals):
        self.ensure_one()
        prev_job_state = self.job_state
        result = super(JobQuoteTemplate, self).write(vals)

        if "job_state" in vals:
            self.handle_job_state(prev_job_state)

        if "approval_status" in vals or "secondary_approval_status" in vals:
            self.set_job_state_is_todo_if_job_quote_is_approved()
            self.set_approved_implement_date_if_approve_job_quote()
            self._create_job_supplier_quotes_if_in_supplier_status()
            self._set_lowest_job_supplier_price_if_in_head_department()
            self.send_email_to_selected_supplier_after_approved()

        if "is_for_crew" in vals:
            self.set_approval_status_selection_based_on_is_for_crew()
            self.approve_quote_if_for_crew()

        if "expected_delivery_date" in vals:
            self._set_deadline_based_on_expected_delivery_date()

        if "implement_date" in vals:
            self.restart_flow_if_changing_invalid_implement_date()
            self._compute_week_number()

        listen.job_quote.call_on_write(self, vals)
        return result

    def unlink(self):

        funcs = listen.job_quote.get_funcs_to_call_after_unlink(self)
        result = super(JobQuoteTemplate, self).unlink()

        for func in funcs:
            func()

        return result

    def action_view_record(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "view_mode": "form",
            "res_id": self.id,
            "target": "current",
            "context": self.env.context,
        }

    def send_email_to_selected_supplier_after_approved(self):
        self.ensure_one()
        if self._is_approved() and not self.is_for_crew and self.job_supplier_quote_id:
            self.job_supplier_quote_id.send_email_to_supplier_to_notify_successed_quote()

    def handle_job_state(self, prev_job_state):
        for record in self:
            message = f"Báo giá chưa được duyệt, vui lòng duyệt báo giá để thực hiện!"

            if not record._is_approved():
                if record.job_state != CONST.UNAPPROVED:
                    raise ValidationError(message)

            else:
                if record.job_state == CONST.UNAPPROVED:
                    message = "Báo giá đã được duyệt, vui lòng chọn trạng thái khác!"
                    raise ValidationError(message)

                elif record.job_state == CONST.COMPLETED:
                    record.finished_at = fields.Date.today()
                    record.material_assignment_ids.end_use(force=True)

                elif record.job_state == CONST.CONFIRMED:
                    record.check_if_user_allow_to_perform_confirm_job_state()
                    record.check_if_material_assignment_have_start_date()
                    record.check_if_material_assignment_have_end_date()
                    if prev_job_state != CONST.COMPLETED:
                        raise ValidationError("Vui lòng hoàn thành báo giá trước!")

    def check_if_material_assignment_have_start_date(self):
        self.ensure_one()
        for material_assignment_id in self.material_assignment_ids:
            ref = material_assignment_id.job_quote_id.ref
            message = f"Nhận vật tư chưa có ngày bắt đầu, vui lòng kiểm tra lại({ref})!"
            if not material_assignment_id.start_time_of_use:
                raise ValidationError(message)

    def check_if_material_assignment_have_end_date(self):
        self.ensure_one()
        for material_assignment_id in self.material_assignment_ids:
            ref = material_assignment_id.job_quote_id.ref
            message = (
                f"Nhận vật tư chưa có ngày kết thúc, vui lòng kiểm tra lại({ref})!"
            )
            if not material_assignment_id.end_time_of_use:
                raise ValidationError(message)

    def check_if_user_allow_to_perform_confirm_job_state(self):
        self.ensure_one()
        group_ids = self._get_groups_for_perform_confirm_button()
        group_xml_ids = []

        for group_id in group_ids:
            if group_id:
                xml_id = list(group_id.ensure_one().get_xml_id().values())[0]
                group_xml_ids.append(xml_id)

        has_group = self.env.user.has_group
        is_user_has_any_group = any([has_group(xml_id) for xml_id in group_xml_ids])

        if not is_user_has_any_group:
            raise ValidationError("Người dùng không có quyền xác nhận báo giá này!")

    def set_approval_status_selection_based_on_is_for_crew(self):
        self.ensure_one()
        if self.is_for_crew:
            self.switch_to_main_approval_status()
        else:
            self.switch_to_secondary_approval_status()

    def approve_quote_if_for_crew(self):
        self.ensure_one()
        if self.is_for_crew:
            self.sudo_approve()

    def set_job_state_is_todo_if_job_quote_is_approved(self):
        for record in self:
            if record._is_approved():
                if record.job_state == CONST.UNAPPROVED:
                    record.bypass_checks().job_state = CONST.TODO
            else:
                if record.job_state != CONST.UNAPPROVED:
                    record.bypass_checks().job_state = CONST.UNAPPROVED

    def set_approved_implement_date_if_approve_job_quote(self):
        self.ensure_one()
        if self._is_approved():
            self.approved_implement_date = self.implement_date

    def restart_flow_if_changing_invalid_implement_date(self):
        self.ensure_one()
        if self._is_approved():
            if not self._is_valid_implement_day():
                self.restart_flow()

    def _is_valid_implement_day(self):
        day_number = self.get_day_number_for_changing_valid_implement_date()

        if self.implement_date < self.approved_implement_date:
            days_difference = (self.approved_implement_date - self.implement_date).days
            return days_difference <= day_number

        if self.implement_date > self.approved_implement_date:
            days_difference = (self.implement_date - self.approved_implement_date).days
            return days_difference <= day_number

    def _confirm_proposal(self):
        self.ensure_one()
        self.job_state = CONST.CONFIRMED

    def _create_job_supplier_quotes_if_in_supplier_status(self):
        self.ensure_one()
        if self.is_at_this_approval_level(CONST.SUPPLIER):
            if not self.is_for_crew:
                if not self.job_supplier_quote_ids:
                    self._create_job_supplier_quotes()

    def _create_job_supplier_quotes(self):
        self.ensure_one()
        supplier_ids = self.job_id.supplier_ids
        for supplier in supplier_ids:
            self.job_supplier_quote_ids.create(
                {
                    "labor_cost": 0,
                    "job_quote_id": self.id,
                    "supplier_id": supplier.id,
                }
            )

    def _set_lowest_job_supplier_price_if_in_head_department(self):
        self.ensure_one()
        if self.is_at_this_approval_level(CONST.HEAD_DEPARTMENT):
            if not self.is_for_crew:
                self._set_lowest_job_supplier_price()

    def _set_lowest_job_supplier_price(self):
        self.ensure_one()
        lowest_price = None
        lowest_supplier_quote = None

        for supplier_quote in self.job_supplier_quote_ids:
            supplier_price = supplier_quote.labor_cost
            if lowest_price is None or supplier_price < lowest_price:
                if supplier_price != 0:
                    lowest_price = supplier_quote.labor_cost
                    lowest_supplier_quote = supplier_quote

        self.bypass_checks().job_supplier_quote_id = lowest_supplier_quote

    def get_the_last_3_times_of_this_quote_for_the_job(self):
        self.ensure_one()
        last_3_times = self.search(
            [
                ("job_id", "=", self.job_id.id),
                ("is_for_crew", "=", False),
                ("implement_date", "<=", self.implement_date),
                ("id", "!=", self.id),
            ],
            limit=3,
        )

        return last_3_times

    def _compute_week_number(self):
        for record in self:
            if record.implement_date:
                week_number = record._get_week_number(record.implement_date)
                record.week_number = week_number
            else:
                not_valid_week_number = 100
                record.week_number = not_valid_week_number

    def _set_deadline_based_on_expected_delivery_date(self):
        self.ensure_one()
        if self.expected_delivery_date:
            deadline_days = timedelta(days=self._get_deadline_days())
            deadline = self.expected_delivery_date - deadline_days
            self.deadline = deadline

    def _get_groups_for_perform_confirm_button(self):
        default_value_model = self._get_default_value_model()
        variable_name = CONST_UTILITIES.GROUPS_SHIP_JOB_QUOTE_TEMPLATE_CHANGING_CONFIRM_JOB_STATE
        return default_value_model._get_default_value_by_variable_name(variable_name)

    def get_day_number_for_changing_valid_implement_date(self):
        default_value_model = self._get_default_value_model()
        variable_name = (
            CONST_UTILITIES.INTEGER_SHIP_JOB_QUOTE_TEMPLATE_DAY_COUNT_VALID_FOT_EDITING_IMPLEMENT_DATE
        )
        return default_value_model._get_default_value_by_variable_name(variable_name)

    def _get_deadline_days(self):
        default_value_model = self._get_default_value_model()
        variable_name = CONST_UTILITIES.INTEGER_SHIP_JOB_QUOTE_TEMPLATE_DEADLINE_DAY_COUNT
        return default_value_model._get_default_value_by_variable_name(variable_name)

    def propose_when_all_suppliers_have_priced(self):
        self.ensure_one()
        labor_costs = self.job_supplier_quote_ids.mapped("labor_cost")

        if all(labor_costs):
            if self.is_at_this_approval_level(CONST.SUPPLIER):
                self.action_propose()


class JobQuote(models.Model):
    _name = "ship.job.quote"
    _description = "Báo giá công việc"
    _inherit = ["ship.job.quote.template"]
    _check_company_auto = True

    _order = "implement_date ASC"

    # related fields
    maintenance_scope_name = fields.Char(
        related="job_id.maintenance_scope_id.name",
        string="Maintnance scope",
        store=False,
    )
    name_for_noti = fields.Char(
        related="job_id.name",
        string="Job",
    )
    labor_cost = fields.Float(
        "Labor cost", related="job_supplier_quote_id.labor_cost", tracking=True
    )
    inspection_date = fields.Date(
        "Inspection Date",
        related="maintenance_scope_report_id.inspection_date",
        store=True,
        readonly=True,
        tracking=True,
        depends=[
            "maintenance_scope_report_id",
            "maintenance_scope_report_id.inspection_date",
        ],
    )

    # sequential
    ref = fields.Char(string="Code", default=lambda self: _("New"))

    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    # relations
    maintenance_scope_report_id = fields.Many2one(
        "ship.maintenance.scope.report",
        string="Maintenance scope report",
        tracking=True,
    )
    maintenance_scope_job_ids = fields.One2many(
        "ship.job",
        "job_quote_id",
        related="maintenance_scope_report_id.maintenance_scope_id.job_ids",
        string="Job",
        tracking=True,
    )
    job_id = fields.Many2one(
        "ship.job",
        string="Job",
        domain="[('id', 'in', maintenance_scope_job_ids)]",
        tracking=True,
    )
    job_supplier_quote_id = fields.Many2one(
        "ship.job.supplier.quote",
        domain="[('job_quote_id', '=', id)]",
        string="selected job supplier",
        tracking=True,
    )
    job_supplier_quote_ids = fields.One2many(
        "ship.job.supplier.quote",
        "job_quote_id",
        string="Job supplier quote",
        readonly=True,
        tracking=True,
    )
    material_assignment_ids = fields.One2many(
        "ship.material.assignment",
        "job_quote_id",
        string="Material Assignment",
        tracking=True,
    )
    area_of_paint_job_ids = fields.One2many(
        "ship.area.of.paint.job",
        "job_quote_id",
        string="Area Of Paint Job",
        tracking=True,
    )

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            vals["ref"] = self.env["ir.sequence"].next_by_code("ship.job.quote")

        result = super(JobQuote, self).create(vals_list)

        for record in result:
            record._set_requirements_if_quote_is_approved()
            record.set_default_material_assignment_start_date_if_quote_is_approved()

        return result

    def write(self, vals):
        self.ensure_one()
        result = super(JobQuote, self).write(vals)

        if "approval_status" in vals or "secondary_approval_status" in vals:
            self._set_requirements_if_quote_is_approved()
            self.set_default_material_assignment_start_date_if_quote_is_approved()

        if "material_assignment_ids" in vals:
            self.set_default_material_assignment_start_date_if_quote_is_approved()

        return result

    def unlink(self):
        for record in self:
            record.material_assignment_ids.unlink()
            record.area_of_paint_job_ids.unlink()

        result = super(JobQuote, self).unlink()

        return result

    def name_get(self):
        result = []
        for report in self:
            name = report.job_id.name or _("New")
            result.append((report.id, name))
        return result

    def set_default_material_assignment_start_date_if_quote_is_approved(self):
        self.ensure_one()
        if self._is_approved() and self.is_for_crew:
            self.material_assignment_ids.start_use(
                start_date=self.implement_date, force=True
            )

    def _set_requirements_if_quote_is_approved(self):
        self.ensure_one()
        if self._is_approved() and self.is_for_crew:
            if not self.material_assignment_ids:
                self._create_material_assignment_based_on_requirement()
            if not self.area_of_paint_job_ids:
                self._create_area_of_paint_based_on_requirement()
        else:
            self.material_assignment_ids.unlink()
            self.area_of_paint_job_ids.unlink()

    def _create_material_assignment_based_on_requirement(self):
        self.ensure_one()
        for requirement in self.job_id.job_material_requirement_ids:
            material_id = requirement.material_id
            required_quantity = requirement.required_quantity

            material_id._assign_materials_based_on_type(required_quantity, self)

    def _create_area_of_paint_based_on_requirement(self):
        assigned_area_of_paints = []
        for requirement in self.job_id.job_paint_requirement_ids:
            assigned_area_of_paints.append(
                {
                    "paint_area_m2": 0,
                    "job_paint_requirement_id": requirement.id,
                    "job_quote_id": self.id,
                }
            )
        self.area_of_paint_job_ids.create(assigned_area_of_paints)

    @api.model
    def export_filtered_job_quotes(self, domain=None):
        # Load your template
        # if domain is None:
        #     domain = []
        template_path = "/mnt/extra-addons/bgt.xlsx"
        workbook = load_workbook(filename=template_path)
        sheet = workbook.active

        # Assuming you want to start writing data from row 2
        row = 2

        # if self.env.context.get('active_ids'):
        active_ids = self.env.context.get("active_ids", [])
        # domain = [('id', 'in', active_ids)] if active_ids else []
        # job_quotes = self.search([('id', 'in', active_ids)])
        job_quotes = self.browse(active_ids)
        for job_quote in job_quotes:
            # Write data to cells
            sheet.cell(row=row, column=1, value=job_quote.ref or "")
            sheet.cell(row=row, column=2, value=job_quote.deadline or "")
            sheet.cell(row=row, column=3, value=job_quote.note or "")
            # ... (write other fields)
            row += 1

        # Save the filled template to a BytesIO object
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Encode the bytes to base64
        xls_content_base64 = base64.b64encode(output.read())
        output.close()

        # Create an attachment
        attachment = self.env["ir.attachment"].create(
            {
                "name": "JobQuotesExport.xlsx",
                "type": "binary",
                "datas": xls_content_base64,
                "store_fname": "JobQuotesExport.xlsx",
                "res_model": "ship.job.quote",
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )

        # Return the action to download the attachment
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "new",
        }
