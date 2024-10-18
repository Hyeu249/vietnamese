from odoo import models
import copy
from odoo.exceptions import ValidationError
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
import tempfile
from openpyxl.styles import Border, Side
from datetime import datetime, timedelta
from odoo import _, api, fields, models
from openpyxl.styles import Border, Side, Alignment


class MaintenanceScopeReportXlsx(models.AbstractModel):
    _name = "report.ship_management.ship_maintenance_scope_report_xlsx"
    _inherit = ["report.report_xlsx.abstract"]

    def generate_xlsx_report(self, _workbook, _data, report_ids):
        for report_id in report_ids:
            self.custom_export_pms_form_to_xlsx(report_id)

    def custom_export_pms_form_to_xlsx(self, report_id):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = (
            "/mnt/extra-addons/report_template/pms.xlsx"  # Replace with the actual path
        )
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise ValidationError(f"Error loading the template: {str(e)}")

        ###

        align_center = Alignment(horizontal="center", vertical="center")
        side_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
        )
        # date_object = fields.Date.from_string(self.proposed_date)
        # day = date_object.day
        # month = date_object.month
        # year = date_object.year
        ### Tên tàu
        worksheet_name = ["Boong", "Máy"]
        for ws in worksheet_name:
            worksheet = workbook[ws]
            worksheet.cell(
                row=3, column=2, value=f"M/V (Tên tàu): {report_id.company_id.name}"
            )
            worksheet[f"B3"].alignment = align_center

        ####PMS
        pms_data_machinery = []  # Initialize lists for different departments
        pms_data_boong = []

        pms_data = report_id.generate_report()

        # Separate data based on department
        for maintainance_data in pms_data:
            if maintainance_data["maintenance_scope_department"] == "MACHINERY":
                pms_data_machinery.append(maintainance_data)
            elif maintainance_data["maintenance_scope_department"] == "BOONG":
                pms_data_boong.append(maintainance_data)

        for pms_data_dep, department_name in [
            (pms_data_machinery, "Máy"),
            (pms_data_boong, "Boong"),
        ]:
            row = 6
            index = 0
            worksheet = workbook[department_name]
            for maintainance_data in pms_data_dep:
                worksheet.insert_rows(row)
                index += 1
                worksheet.cell(row=row, column=1, value=index)
                worksheet.cell(
                    row=row,
                    column=2,
                    value=maintainance_data["equipment_name"]
                    + "-"
                    + maintainance_data["maintenance_scope_name"],
                )
                worksheet.cell(
                    row=row,
                    column=3,
                    value=maintainance_data["job_name"],
                )
                worksheet.cell(
                    row=row,
                    column=4,
                    value=f"{int(int(maintainance_data['maintenance_scope_maintenance_interval_days'])/30)} tháng",
                )

                date_object = fields.Date.from_string(
                    maintainance_data["maintenance_scope_last_maintenance_date"]
                )
                day = date_object.day
                month = date_object.month
                year = date_object.year

                worksheet.cell(
                    row=row,
                    column=5,
                    value=f"{maintainance_data['maintenance_scope_last_maintenance_date']}",
                )

                date_object = fields.Date.from_string(
                    maintainance_data["maintenance_scope_last_maintenance_date"]
                )
                day = date_object.day
                month = date_object.month
                year = date_object.year
                end_of_year = datetime(2024, 12, 31)
                # Create a datetime object from the extracted components
                last_maintenance_date = datetime(year, month, day)
                current_date = last_maintenance_date
                expected_dates = []
                while current_date < end_of_year:
                    expected_dates.append(current_date)
                    current_date += timedelta(
                        days=int(
                            maintainance_data[
                                "maintenance_scope_maintenance_interval_days"
                            ]
                        )
                    )
                for date in expected_dates:
                    month_number = date.month
                    year = date.year
                    cell = worksheet.cell(
                        row=row,
                        column=month_number + 6,
                        value="Y",
                    )
                row += 1

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            border_style = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            worksheet_name = ["Boong", "Máy"]
            for ws in worksheet_name:
                worksheet = workbook[ws]
                for row in range(6, index + 1):
                    for col in range(1, 19):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = border_style

            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Kế hoạch PMS.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": report_id._name,
                "res_id": report_id.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action
