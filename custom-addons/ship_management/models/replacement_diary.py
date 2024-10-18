# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from . import CONST
from odoo.exceptions import ValidationError
from datetime import date
from datetime import timedelta
from .common_utils import generate_token, format_field_date
import logging
import pandas as pd
import io
import os
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import tempfile
from openpyxl.styles import Border, Side

from odoo.exceptions import UserError
from odoo import _, api, fields, models
from datetime import datetime, timedelta
from openpyxl.styles import Border, Side, Alignment


class ReplacementDiary(models.Model):
    _name = "ship.replacement.diary"
    _description = "Replacement diary records"
    _inherit = ["utilities.notification"]
    _check_company_auto = True

    date = fields.Date("Date", tracking=True)
    description = fields.Char("Description", tracking=True)
    quantity = fields.Float("Quantity", tracking=True)
    reason = fields.Char("Reason", tracking=True)
    condition = fields.Char("Condition", tracking=True)
    year_used = fields.Char("Year Used", tracking=True)
    note = fields.Char("Note", tracking=True)
    internal_code = fields.Char("Material Internal code", tracking=True)
    material_description = fields.Char("Material Description", tracking=True)
    is_material_visible = fields.Boolean(
        "Is Material Visible", compute="_compute_is_material_visible"
    )
    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    # relations
    material_id = fields.Many2one(
        "ship.material",
        string="Material",
        tracking=True,
        domain="[('material_type', '!=', 'CONSUMABLE_MATERIAL')]",
    )

    material_entity_id = fields.Many2one(
        "ship.material.entity",
        string="Material Entity",
        tracking=True,
        domain="[('material_id.material_type', '!=', 'CONSUMABLE_MATERIAL')]",
    )

    proposed_liquidation_id = fields.Many2one(
        "ship.proposed.liquidation", string="Proposal Liquidation", tracking=True
    )

    # sequential
    ref = fields.Char(string="Code", default=lambda self: _("New"))

    @api.onchange("material_id")
    def _onchange_material_id(self):
        if self.material_id:
            self.internal_code = self.material_id.internal_code
            self.material_description = self.material_id.description
        else:
            self.internal_code = False
            self.material_description = False

    @api.onchange("material_entity_id")
    def _onchange_material_entity_id(self):
        if self.material_entity_id:
            self.material_id = self.material_entity_id.material_id
            self.internal_code = self.material_id.internal_code
            self.material_description = f"(Vật tư thực tế {self.material_entity_id.ref}) {self.material_id.description}"
            self.quantity = 1
        else:
            self.internal_code = False
            self.material_description = False

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            vals["ref"] = self.env["ir.sequence"].next_by_code("ship.replacement.diary")
        return super(ReplacementDiary, self).create(vals_list)

    def name_get(self):
        result = []
        for report in self:
            name = report.ref or _("New")
            result.append((report.id, name))
        return result

    @api.depends("material_entity_id")
    def _compute_is_material_visible(self):
        for record in self:
            record.is_material_visible = not record.material_entity_id

    def generate_report(self):
        diary_data = []
        replacement_diaries = self.search([])  # Fetch all replacement diary records

        for diary in replacement_diaries:
            diary_data.append(
                {
                    "material_name": (
                        diary.material_id.name if diary.material_id else ""
                    ),
                    "replacement_diary_date": diary.date,
                    "replacement_diary_description": diary.description,
                    "replacement_diary_quantity": diary.quantity,
                    "replacement_diary_reason": diary.reason,
                    "replacement_diary_condition": diary.condition,
                    "replacement_diary_year_used": diary.year_used,
                    "replacement_diary_note": diary.note,
                    "replacement_diary_internal_code": diary.internal_code,
                    "replacement_diary_material_description": diary.material_description,
                    "replacement_diary_is_material_visible": diary.is_material_visible,
                    "replacement_diary_company_id": diary.company_id.name,
                    "replacement_diary_material_id": diary.material_id.id,
                    "replacement_diary_material_entity_id": diary.material_entity_id.id,
                    "replacement_diary_proposed_liquidation_id": diary.proposed_liquidation_id.id,
                    "replacement_diary_unit": (
                        diary.material_id.material_type if diary.material_id else ""
                    ),
                    "material_warehouse": (
                        diary.material_id.warehouse if diary.material_id else ""
                    ),
                    "proposed_date": (
                        diary.proposed_liquidation_id.proposed_date
                        if diary.proposed_liquidation_id
                        else ""
                    ),
                }
            )

        return diary_data

    def custom_export_diary_to_xlsx(self):
        # Load the custom XLSX template
        template_path = "/mnt/extra-addons/report_template/nktr.xlsx"  # Replace with the actual path
        try:
            workbook = load_workbook(template_path)
            worksheet_names = workbook.sheetnames
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        # Function to adjust merged cells after inserting rows
        align_center = Alignment(horizontal="center", vertical="center")
        side_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
        )

        # Initialize lists for different departments
        diary_data = self.generate_report()

        index_m = 0
        index_b = 0
        index_t = 0
        for diary in diary_data:
            material_warehouse = diary["material_warehouse"]
            if material_warehouse == "machinery":
                ws = "Máy"
                worksheet = workbook[ws]
                worksheet.cell(
                    row=1, column=1, value=f"Tên tàu: {self.company_id.name}"
                )
                worksheet.cell(row=1, column=8, value=f"Bộ phận: {ws}")
                index_m += 1
                current_row = 2 + index_m
                worksheet.cell(row=current_row, column=1, value=index_m)
            elif material_warehouse == "boong":
                ws = "Boong"
                worksheet = workbook[ws]
                worksheet.cell(
                    row=1, column=1, value=f"Tên tàu: {self.company_id.name}"
                )
                worksheet.cell(row=1, column=8, value=f"Bộ phận: {ws}")
                index_b += 1
                current_row = 2 + index_b
                worksheet.cell(row=current_row, column=1, value=index_b)
            elif material_warehouse == "tool":
                ws = "Dụng cụ"
                worksheet = workbook[ws]
                worksheet.cell(
                    row=1, column=1, value=f"Tên tàu: {self.company_id.name}"
                )
                worksheet.cell(row=1, column=8, value=f"Bộ phận: {ws}")
                index_t += 1
                current_row = 2 + index_t
                worksheet.cell(row=current_row, column=1, value=index_t)

            # Add the entry to the corresponding worksheet
            worksheet.cell(row=current_row, column=2, value=f"{diary['material_name']}")
            worksheet.cell(
                row=current_row, column=3, value=diary["replacement_diary_description"]
            )
            worksheet.cell(
                row=current_row,
                column=4,
                value=diary["replacement_diary_internal_code"],
            )
            worksheet.cell(
                row=current_row, column=5, value=diary["replacement_diary_unit"]
            )
            worksheet.cell(
                row=current_row, column=6, value=diary["replacement_diary_quantity"]
            )
            worksheet.cell(
                row=current_row, column=7, value=diary["replacement_diary_reason"]
            )
            worksheet.cell(
                row=current_row, column=8, value=diary["replacement_diary_date"]
            )
            worksheet.cell(
                row=current_row, column=9, value=diary["replacement_diary_year_used"]
            )
            worksheet.cell(
                row=current_row, column=10, value=diary["replacement_diary_condition"]
            )
            worksheet.cell(
                row=current_row, column=11, value=diary["replacement_diary_note"]
            )
            worksheet.cell(row=current_row, column=12, value=diary["proposed_date"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified workbook to the temporary file
            workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Replacement Diary.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export_diary_to_xlsx",
            }
        )

        # Create an action to open the attachment
        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        return action
