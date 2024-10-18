# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from . import CONST
from odoo.exceptions import ValidationError
import io
import os
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import tempfile
from odoo.exceptions import UserError
from odoo import _, api, fields, models
from datetime import datetime, timedelta
from openpyxl.styles import Alignment


class CostSettlementReport(models.Model):
    _name = "docking.cost.settlement.report"
    _description = "Quyết toán chi phí"
    _inherit = ["utilities.approval.status", "utilities.notification"]
    _check_company_auto = True

    comment = fields.Char("Comment", tracking=True)
    is_matched_with_supplier = fields.Boolean("Is matched with supplier", tracking=True)

    # related
    name_for_noti = fields.Char(
        related="docking_plan_id.name",
        string="Docking name",
    )

    # relations
    docking_plan_id = fields.Many2one(
        "docking.docking.plan",
        string="Docking plan",
        tracking=True,
    )
    supplier_id = fields.Many2one(
        "docking.supplier",
        string="Supplier",
        tracking=True,
    )
    material_quote_ids = fields.One2many(
        "docking.material.quote",
        "cost_settlement_report_id",
        string="Material Quote",
        # readonly=True,
        tracking=True,
    )
    job_quote_ids = fields.One2many(
        "docking.job.quote",
        "cost_settlement_report_id",
        string="Job Quote",
        # readonly=True,
        tracking=True,
    )
    is_confirm_is_matched_with_supplier_btn_visible = fields.Boolean(
        "Is confirm is matched with supplier btn visible",
        compute="_get_is_confirm_is_matched_with_supplier_btn_visible",
    )

    # Define SQL constraint
    _sql_constraints = [
        (
            "unique_docking_plan_id_supplier_id",
            "unique (docking_plan_id, supplier_id)",
            "docking_plan_id and supplier_id must be unique.",
        ),
    ]

    # sequential
    ref = fields.Char(string="Code", default=lambda self: _("New"))

    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            model_name = "docking.cost.settlement.report"
            vals["ref"] = self.env["ir.sequence"].next_by_code(model_name)

        result = super(CostSettlementReport, self).create(vals_list)

        return result

    def name_get(self):
        result = []
        for report in self:
            name = report.ref or _("New")
            result.append((report.id, name))
        return result

    def get_approved_material_job_quote_ids_by_supplier(self):
        self.ensure_one()
        material_quote_ids = self._get_approved_material_quote_ids_by_supplier()
        job_quote_ids = self._get_approved_job_quote_ids_by_supplier()

        self.material_quote_ids = [(6, 0, material_quote_ids)]
        self.job_quote_ids = [(6, 0, job_quote_ids)]

    def _get_approved_material_quote_ids_by_supplier(self):
        material_survey_data_ids = self.docking_plan_id.material_survey_data_ids
        material_quote_ids = [
            quote.id
            for survey in material_survey_data_ids
            for quote in survey.material_quote_ids
            if quote.material_supplier_quote_id.supplier_id == self.supplier_id
            and quote._is_approved()
        ]

        return material_quote_ids

    def _get_approved_job_quote_ids_by_supplier(self):
        equipment_survey_data_ids = self.docking_plan_id.equipment_survey_data_ids
        job_quotes = [
            quote
            for survey in equipment_survey_data_ids
            for report in survey.maintenance_scope_report_ids
            for quote in report.job_quote_ids
            if quote.job_supplier_quote_id.supplier_id == self.supplier_id
            and quote._is_approved()
        ]

        job_quote_ids = [quote.id for quote in job_quotes]
        return job_quote_ids

    def custom_export_job_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = "/mnt/extra-addons/report_template/hmsc.xlsx"  # Replace with the actual path
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###
        row = 3
        index = 0
        for job_quote in self.job_quote_ids:
            worksheet.insert_rows(row)
            index += 1
            worksheet.cell(row=row, column=1, value=index)
            worksheet.cell(
                row=row,
                column=2,
                value=job_quote.maintenance_scope_report_id.equipment_survey_data_id.equipment_survey_metadata_id.name,
            )
            worksheet.cell(
                row=row,
                column=3,
                value=job_quote.maintenance_scope_report_id.maintenance_scope_id.name,
            )
            worksheet.cell(row=row, column=4, value=job_quote.name_for_noti)
            worksheet.cell(row=row, column=5, value=job_quote.specification)
            worksheet.cell(row=row, column=6, value=job_quote.length)
            worksheet.cell(row=row, column=7, value=job_quote.width)
            worksheet.cell(row=row, column=8, value=job_quote.height)
            worksheet.cell(row=row, column=9, value=job_quote.quantity)
            worksheet.cell(row=row, column=10, value=job_quote.unit)
            worksheet.cell(row=row, column=11, value=job_quote.note)
            worksheet.cell(row=row, column=12, value=job_quote.labor_cost)
            worksheet.cell(row=row, column=13, value=job_quote.factor)
            worksheet.cell(row=row, column=14, value=job_quote.final_cost)
            row += 1

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Quyết toán công việc.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action

    def custom_export_material_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = (
            "/mnt/extra-addons/report_template/vt.xlsx"  # Replace with the actual path
        )
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###
        row = 3
        index = 0
        for material_quote in self.material_quote_ids:
            worksheet.insert_rows(row)
            index += 1
            worksheet.cell(row=row, column=1, value=index)
            worksheet.cell(
                row=row,
                column=2,
                value=material_quote.material_survey_data_id.material_survey_metadata_id.description,
            )
            worksheet.cell(
                row=row,
                column=3,
                value=material_quote.material_survey_data_id.material_survey_metadata_id.spare_part_no,
            )
            worksheet.cell(row=row, column=4, value=material_quote.material_name)
            worksheet.cell(
                row=row,
                column=5,
                value=material_quote.material_survey_data_id.material_survey_metadata_id.unit,
            )
            worksheet.cell(row=row, column=7, value=material_quote.quantity)
            worksheet.cell(row=row, column=8, value=material_quote.unit_price)
            worksheet.cell(
                row=row,
                column=9,
                value=(material_quote.unit_price * material_quote.quantity),
            )
            worksheet.cell(row=row, column=10, value=material_quote.note)
            row += 1

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Quyết toán vật tư.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action

    def action_propose(self):
        # if it is about to be proposed to the last approval level,
        # check if it is matched with supplier
        if self.is_second_last_approval_level():
            if not self.is_matched_with_supplier:
                raise ValidationError(
                    "Báo cáo chưa được xác nhận là khớp với nhà cung cấp, vui lòng kiểm tra lại!"
                )

        return super(CostSettlementReport, self).action_propose()

    def action_confirm_is_matched_with_supplier(self):
        self.ensure_one()
        self.is_matched_with_supplier = True

    def action_confirm_is_not_matched_with_supplier(self):
        self.ensure_one()
        self.is_matched_with_supplier = False
        # reset approval flow
        self.restart_flow()

    @api.onchange("approval_status")
    def _get_is_confirm_is_matched_with_supplier_btn_visible(self):
        self.ensure_one()
        if self._get_group_id_based_on_approval_status() is False:
            self.is_confirm_is_matched_with_supplier_btn_visible = False
            return
        if self.is_second_last_approval_level():
            self.is_confirm_is_matched_with_supplier_btn_visible = True
            return
        self.is_confirm_is_matched_with_supplier_btn_visible = False
