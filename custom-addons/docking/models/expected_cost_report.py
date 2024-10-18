# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from . import CONST
from odoo.exceptions import ValidationError
import io
import os
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import tempfile
from odoo.exceptions import UserError
from odoo import _, api, fields, models
from datetime import datetime, timedelta
from openpyxl.styles import Alignment


class ExpectedCostReport(models.Model):
    _name = "docking.expected.cost.report"
    _description = "Dự toán chi phí"
    _inherit = ["utilities.approval.status", "mail.thread"]
    _check_company_auto = True

    comment = fields.Char("Comment", tracking=True)
    material_total_price = fields.Char(
        "Material total price", compute="_get_material_total_price", tracking=True
    )
    job_total_price = fields.Char(
        "Job total price", compute="_get_job_total_price", tracking=True
    )

    name_for_noti = fields.Char(
        related="docking_plan_id.name",
        string="Docking name",
    )

    # relations
    docking_plan_id = fields.Many2one(
        "docking.docking.plan",
        string="Docking plan",
        tracking=True,
    )
    material_quote_ids = fields.One2many(
        "docking.material.quote",
        "expected_cost_report_id",
        string="Material Quote",
        domain="[('material_survey_data_id.docking_plan_id', '=', docking_plan_id)]",
    )
    job_quote_ids = fields.One2many(
        "docking.job.quote",
        "expected_cost_report_id",
        string="Job Quotes",
        domain="[('maintenance_scope_report_id.equipment_survey_data_id.docking_plan_id', '=', docking_plan_id)]",
    )

    # Define SQL constraint
    _sql_constraints = [
        (
            "unique_docking_plan_id",
            "unique (docking_plan_id)",
            "docking_plan_id must be unique.",
        ),
    ]

    # sequential
    ref = fields.Char(string="Code", default=lambda self: _("New"))

    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            model_name = "docking.expected.cost.report"
            vals["ref"] = self.env["ir.sequence"].next_by_code(model_name)

        results = super(ExpectedCostReport, self).create(vals_list)
        return results

    def write(self, vals):
        old_job_quote_ids = self.job_quote_ids.ids
        old_material_q_ids = self.material_quote_ids.ids
        old_job_quotes = [
            {quote.id: quote.name_for_noti} for quote in self.job_quote_ids
        ]
        old_material_qs = [
            {quote.id: quote.name_for_noti} for quote in self.material_quote_ids
        ]

        result = super(ExpectedCostReport, self).write(vals)
        new_job_quote_ids = self.job_quote_ids.ids
        new_material_q_ids = self.material_quote_ids.ids

        if old_job_quote_ids != new_job_quote_ids:
            added_quotes = list(set(new_job_quote_ids) - set(old_job_quote_ids))
            removed_quotes = list(set(old_job_quote_ids) - set(new_job_quote_ids))
            removed_names = self.get_removed_quote_name(removed_quotes, old_job_quotes)

            self._log_quote_changes_to_chatter(added_quotes, removed_names)

        if old_material_q_ids != new_material_q_ids:
            added_quotes = list(set(new_material_q_ids) - set(old_material_q_ids))
            removed_quotes = list(set(old_material_q_ids) - set(new_material_q_ids))
            removed_names = self.get_removed_quote_name(removed_quotes, old_material_qs)

            self._log_material_quote_changes_to_chatter(added_quotes, removed_names)

        docking_plan_id = self.docking_plan_id
        if "approval_status" in vals:
            are_all_surveys_approved = docking_plan_id._are_all_survey_datas_approved()

            if not are_all_surveys_approved:
                raise ValidationError("Các khảo sát chưa được duyệt xong!")
            if not self._are_all_material_quotes_approved():
                raise ValidationError("Báo giá vật tư chưa được duyệt xong!")
            if not self._are_all_job_quotes_approved():
                raise ValidationError("Báo giá công việc chưa được duyệt xong")

        return result

    def get_removed_quote_name(self, removed_ids, quotes):
        removed_name = []

        for quote in quotes:
            for key in quote:
                if key in removed_ids:
                    removed_name.append(quote[key])

    def _log_quote_changes_to_chatter(self, added_quotes, removed_names):
        message = ""
        if added_quotes:
            added_names = (
                self.env["docking.job.quote"]
                .browse(added_quotes)
                .mapped("name_for_noti")
            )
            message += "<li>Added quotes: {}</li>".format(", ".join(added_names))

        if removed_names:
            message += "<li>Removed quotes: {}</li>".format(", ".join(removed_names))

        if message:
            full_message = "<ul>{}</ul>".format(message)
            self.message_post(body=full_message)

    def _log_material_quote_changes_to_chatter(
        self, added_material_quotes, removed_names
    ):
        message = ""
        if added_material_quotes:
            added_names = (
                self.env["docking.material.quote"]
                .browse(added_material_quotes)
                .mapped("name_for_noti")
            )
            message += "<li>Added material quotes: {}</li>".format(
                ", ".join(added_names)
            )

        if removed_names:
            message += "<li>Removed material quotes: {}</li>".format(
                ", ".join(removed_names)
            )
        if message:
            full_message = "<ul>{}</ul>".format(message)
            self.message_post(body=full_message)

    def name_get(self):
        result = []
        for report in self:
            name = report.ref or _("New")
            result.append((report.id, name))
        return result

    def _are_all_material_quotes_approved(self):
        self.ensure_one()

        all_quotes_approved = all([a._is_approved() for a in self.material_quote_ids])
        return all_quotes_approved

    def _are_all_job_quotes_approved(self):
        self.ensure_one()

        all_quotes_approved = all([a._is_approved() for a in self.job_quote_ids])
        return all_quotes_approved

    def _get_supplier_ids(self):
        material_supplier_ids = self._get_supplier_ids_from_material_quotes()
        job_supplier_ids = self._get_supplier_ids_from_job_quotes()

        supplier_ids = material_supplier_ids + job_supplier_ids
        return list(set(supplier_ids))

    def _get_supplier_ids_from_material_quotes(self):
        supplier_ids = [
            quote.material_supplier_quote_id.supplier_id.id
            for quote in self.material_quote_ids
            if quote.material_supplier_quote_id
        ]
        return supplier_ids

    def _get_supplier_ids_from_job_quotes(self):
        supplier_ids = [
            quote.job_supplier_quote_id.supplier_id.id
            for quote in self.job_quote_ids
            if quote.job_supplier_quote_id
        ]
        return supplier_ids

    def create_contract(self):
        self.ensure_one()
        docking_plan_id = self.docking_plan_id

        if not self._is_approved():
            raise ValidationError("Chưa duyệt dự toán chi phí!")
        if docking_plan_id.contract_ids:
            raise ValidationError("Đã tạo các hợp đồng!")
        if not docking_plan_id.start_date:
            raise ValidationError("Chưa có ngày docking thực tế!")

        for supplier_id in self._get_supplier_ids():
            self.env["docking.contract"].create(
                {
                    "docking_plan_id": docking_plan_id.id,
                    "supplier_id": supplier_id,
                }
            )

    def get_new_material_job_quote_ids(self):
        self.ensure_one()
        material_quote_ids = self._get_material_quote_ids()
        job_quote_ids = self._get_job_quote_ids()

        self.material_quote_ids = [(6, 0, material_quote_ids)]
        self.job_quote_ids = [(6, 0, job_quote_ids)]

    def _get_material_quote_ids(self):
        surveys = self.docking_plan_id.material_survey_data_ids
        material_survey_data_ids = surveys.filtered(lambda e: not e._is_arise())
        material_quote_ids = [
            quote.id
            for survey in material_survey_data_ids
            for quote in survey.material_quote_ids
        ]
        return material_quote_ids

    def _get_job_quote_ids(self):
        surveys = self.docking_plan_id.equipment_survey_data_ids
        equipment_survey_data_ids = surveys.filtered(lambda e: not e._is_arise())
        job_quote_ids = [
            quote.id
            for survey in equipment_survey_data_ids
            for report in survey.maintenance_scope_report_ids
            for quote in report.job_quote_ids
            if not quote.is_for_crew
        ]
        return job_quote_ids

    @api.depends("material_quote_ids")
    def _get_material_total_price(self):
        for record in self:
            quotes = self.material_quote_ids
            record.material_total_price = sum([quote.total_price for quote in quotes])

    @api.depends("job_quote_ids")
    def _get_job_total_price(self):
        for record in self:
            quotes = self.job_quote_ids
            record.job_total_price = sum([quote.final_cost for quote in quotes])

    def custom_export_job_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = "/mnt/extra-addons/report_template/hmsc.xlsx"  # Replace with the actual path
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###
        row = 3
        index = 0
        for job_quote in self.job_quote_ids:
            worksheet.insert_rows(row)
            index += 1
            worksheet.cell(row=row, column=1, value=index)
            worksheet.cell(
                row=row,
                column=2,
                value=job_quote.maintenance_scope_report_id.equipment_survey_data_id.equipment_survey_metadata_id.name,
            )
            worksheet.cell(
                row=row,
                column=3,
                value=job_quote.maintenance_scope_report_id.maintenance_scope_id.name,
            )
            worksheet.cell(row=row, column=4, value=job_quote.name_for_noti)
            worksheet.cell(row=row, column=5, value=job_quote.specification)
            worksheet.cell(row=row, column=6, value=job_quote.length)
            worksheet.cell(row=row, column=7, value=job_quote.width)
            worksheet.cell(row=row, column=8, value=job_quote.height)
            worksheet.cell(row=row, column=9, value=job_quote.quantity)
            worksheet.cell(row=row, column=10, value=job_quote.unit)
            worksheet.cell(row=row, column=11, value=job_quote.note)
            worksheet.cell(row=row, column=12, value=job_quote.labor_cost)
            worksheet.cell(row=row, column=13, value=job_quote.factor)
            worksheet.cell(row=row, column=14, value=job_quote.final_cost)
            row += 1

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Dự toán công việc.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action

    def custom_export_material_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = (
            "/mnt/extra-addons/report_template/vt.xlsx"  # Replace with the actual path
        )
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###
        row = 3
        index = 0
        for material_quote in self.material_quote_ids:
            worksheet.insert_rows(row)
            index += 1
            worksheet.cell(row=row, column=1, value=index)
            worksheet.cell(
                row=row,
                column=2,
                value=material_quote.material_survey_data_id.material_survey_metadata_id.description,
            )
            worksheet.cell(
                row=row,
                column=3,
                value=material_quote.material_survey_data_id.material_survey_metadata_id.spare_part_no,
            )
            worksheet.cell(row=row, column=4, value=material_quote.material_name)
            worksheet.cell(
                row=row,
                column=5,
                value=material_quote.material_survey_data_id.material_survey_metadata_id.unit,
            )
            worksheet.cell(row=row, column=7, value=material_quote.quantity)
            worksheet.cell(row=row, column=8, value=material_quote.unit_price)
            worksheet.cell(
                row=row,
                column=9,
                value=(material_quote.unit_price * material_quote.quantity),
            )
            worksheet.cell(row=row, column=10, value=material_quote.note)
            row += 1

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Dự toán vật tư.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action
