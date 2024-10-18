# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from . import CONST
from ...utilities.models import CONST as CONST_UTILITIES
from odoo.exceptions import ValidationError
from datetime import date
from datetime import timedelta
from .common_utils import generate_token, format_field_date
import logging
import pandas as pd
import io
import os
import base64  # Add this import for base64 encoding
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import tempfile
from openpyxl.styles import Border, Side

from odoo.exceptions import UserError
from odoo import _, api, fields, models
from datetime import datetime, timedelta
from openpyxl.styles import Border, Side, Alignment
from . import listen


class MaintenanceScopeReportTemplate(models.Model):
    _name = "ship.maintenance.scope.report.template"
    _description = "Maintenance scope report records"
    _inherit = ["ship.date"]

    rq_number = fields.Char("RQ number", tracking=True)
    problem_description = fields.Char("Problem description", tracking=True)
    ship_location = fields.Char("Ship location", tracking=True)
    reason = fields.Char("Reason", tracking=True)
    time_consumption = fields.Integer("Time consumption", tracking=True)
    request_date = fields.Date("Request date", tracking=True)
    expected_delivery_date = fields.Date("Expected delivery date", tracking=True)
    deadline = fields.Date("Deadline", tracking=True)

    # compute fields
    expected_implement_date = fields.Date("Expected implement date", tracking=True)
    finished_at = fields.Date("Finished at", tracking=True)
    total_price = fields.Float("Total price", compute="_get_total_price", tracking=True)
    not_allowed_to_confirm_all_job_quotes = fields.Boolean("Not allow crew")
    job_quote_approval_statuses = fields.Char(
        string="Job quote approval statuses", compute="set_job_quote_approval_statuses"
    )
    progress_by_date = fields.Selection(
        CONST.SCHEDULE_STATES,
        string="Schedule",
        compute="_set_progress_by_date",
        tracking=True,
    )
    progress = fields.Integer(string="Progress", compute="_compute_progress")
    ref_job_state = fields.Selection(
        "get_ref_job_state",
        string="Job state",
        default=CONST.UNAPPROVED,
        group_expand="_expand_groups",
        readonly=True,
        tracking=True,
    )

    def get_ref_job_state(self):
        if self._name == "ship.maintenance.scope.report":
            return CONST.JOB_STATES_FOR_SHIP
        else:
            return CONST.JOB_STATES

    @api.model
    def _expand_groups(self, states, domain, order):
        if self._name == "ship.maintenance.scope.report":
            return [
                CONST.WAITING_FOR_DOCKING,
                CONST.UNAPPROVED,
                CONST.TODO,
                CONST.WORKING,
                CONST.COMPLETED,
                CONST.CONFIRMED,
            ]
        else:
            return [
                CONST.UNAPPROVED,
                CONST.TODO,
                CONST.WORKING,
                CONST.COMPLETED,
                CONST.CONFIRMED,
            ]

    priority = fields.Selection(
        CONST.PRIORITY_SELECTION,
        string="Priority",
        tracking=True,
    )
    ribbon_color = fields.Integer(
        string="Color", compute="_set_ribbon_color", default=CONST.YELLOW_CODE
    )
    request_state = fields.Selection(
        CONST.JOB_QUOTE_REQUEST_STATES,
        string="Request state",
        default=CONST.PREPARE,
        readonly=True,
        tracking=True,
    )
    is_all_for_crew = fields.Boolean("Is all for crew", compute="_calc_is_all_for_crew")
    uncompleted_jobs = fields.Char("Uncompleted job", compute="_calc_uncompleted_job")
    week_number = fields.Integer(string="Week Number")

    # relations
    maintenance_scope_id = fields.Many2one(
        "ship.maintenance.scope",
        string="Maintenance Scope",
        tracking=True,
    )
    job_quote_ids = fields.One2many(
        "ship.job.quote",
        "maintenance_scope_report_id",
        string="Job quote",
        tracking=True,
    )

    # listen
    @listen.job_quote.on_create
    def job_quote_created_in_report(self):
        for record in self:
            record.set_ref_job_state()
            record.set_started_at()
            record.set_finished_at()

    @listen.job_quote.on_write("job_state")
    def job_quote_writed_in_report(self):
        for record in self:
            record.set_ref_job_state()
            record.finish_report_if_all_job_quote_finished()

    @listen.job_quote.on_unlink
    def job_quote_unlink_in_report(self):
        for record in self:
            record.set_ref_job_state()
            record.set_started_at()
            record.set_finished_at()

    def set_ref_job_state(self):
        for record in self:
            job_quote_ids = record.job_quote_ids
            job_state_arr = job_quote_ids.mapped("job_state")

            all_completed = all([st == CONST.COMPLETED for st in job_state_arr])
            all_confirmed = all([st == CONST.CONFIRMED for st in job_state_arr])
            all_todo = all([st == CONST.TODO for st in job_state_arr])

            unapprove_in_state = CONST.UNAPPROVED in job_state_arr
            todo_in_state = CONST.TODO in job_state_arr
            working_in_state = CONST.WORKING in job_state_arr
            complete_in_state = CONST.COMPLETED in job_state_arr
            confirm_in_state = CONST.CONFIRMED in job_state_arr

            if all_todo:
                record.ref_job_state = CONST.TODO

            elif all_completed:
                record.ref_job_state = CONST.COMPLETED

            elif all_confirmed:
                record.ref_job_state = CONST.CONFIRMED

            elif unapprove_in_state:
                record.ref_job_state = CONST.UNAPPROVED

            elif working_in_state:
                record.ref_job_state = CONST.WORKING

            elif todo_in_state and complete_in_state:
                record.ref_job_state = CONST.WORKING

            elif todo_in_state and confirm_in_state:
                record.ref_job_state = CONST.WORKING

            elif complete_in_state and confirm_in_state:
                record.ref_job_state = CONST.COMPLETED

            else:
                record.ref_job_state = CONST.UNAPPROVED

    @listen.job_quote.on_write("implement_date")
    def set_started_at(self):
        for record in self:
            if self.job_quote_ids:
                implement_date_arr = self.job_quote_ids.mapped("implement_date")

                if any(
                    [implement_date == False for implement_date in implement_date_arr]
                ):
                    record.expected_implement_date = False
                else:
                    implement_date_arr.sort()
                    record.expected_implement_date = implement_date_arr[0]

            else:
                record.expected_implement_date = False

    @listen.job_quote.on_write("finished_at")
    def set_finished_at(self):
        for record in self:
            if self.job_quote_ids:
                finished_at_arr = self.job_quote_ids.mapped("finished_at")

                if any([finished_at == False for finished_at in finished_at_arr]):
                    record.finished_at = False
                else:
                    finished_at_arr.sort()
                    record.finished_at = finished_at_arr[-1]

            else:
                record.finished_at = False

    # computes
    @api.depends("job_quote_ids")
    def set_job_quote_approval_statuses(self):
        for record in self:
            if record.job_quote_ids:
                approved_quotes = record.is_all_quotes_approved()
                group_names = record.get_group_names_of_unapproved_quotes()

                if approved_quotes:
                    record.job_quote_approval_statuses = CONST.APPROVED
                else:
                    record.job_quote_approval_statuses = ", ".join(group_names)
            else:
                record.job_quote_approval_statuses = ""

    @api.depends("expected_implement_date", "ref_job_state")
    def _set_ribbon_color(self):
        for record in self:
            expected_implement_date = record.expected_implement_date
            ref_job_state = record.ref_job_state
            today = fields.Date.today()
            if expected_implement_date:
                if expected_implement_date > today:
                    record.ribbon_color = CONST.GREEN_CODE
                elif expected_implement_date == today:
                    record.ribbon_color = CONST.YELLOW_CODE
                elif expected_implement_date < today:
                    if ref_job_state in [
                        CONST.WORKING,
                        CONST.COMPLETED,
                        CONST.CONFIRMED,
                    ]:
                        record.ribbon_color = None
                    else:
                        record.ribbon_color = CONST.RED_CODE
            else:
                record.ribbon_color = None

    @api.depends("expected_implement_date", "ref_job_state")
    def _set_progress_by_date(self):
        for record in self:
            expected_implement_date = record.expected_implement_date
            ref_job_state = record.ref_job_state
            today = fields.Date.today()
            if expected_implement_date:
                if expected_implement_date > today:
                    record.progress_by_date = CONST.UNDUE
                elif expected_implement_date == today:
                    record.progress_by_date = CONST.TODAY
                elif expected_implement_date < today:
                    if ref_job_state in [
                        CONST.WORKING,
                        CONST.COMPLETED,
                        CONST.CONFIRMED,
                    ]:
                        record.progress_by_date = CONST.BLANK
                    else:
                        record.progress_by_date = CONST.OVERDUE
            else:
                record.progress_by_date = None

    @api.depends("job_quote_ids")
    def _compute_progress(self):
        for record in self:
            if record.job_quote_ids:
                job_states = record.job_quote_ids.mapped("job_state")
                confirm_count = job_states.count(CONST.CONFIRMED)
                total_count = len(job_states)

                record.progress = (confirm_count / total_count) * 100
            else:
                record.progress = 0

    @api.depends("job_quote_ids")
    def _calc_uncompleted_job(self):
        for record in self:
            uncompleted_job_quotes = record.job_quote_ids.filtered(
                lambda jq: jq.job_state != CONST.COMPLETED
            )

            jobs = [str(quote.job_id.name) for quote in uncompleted_job_quotes]
            record.uncompleted_jobs = ", ".join(jobs)

    @api.depends("job_quote_ids")
    def _calc_is_all_for_crew(self):
        for record in self:
            for_crew_arr = record.job_quote_ids.mapped("is_for_crew")
            record.is_all_for_crew = all(for_crew_arr)

    @api.depends("job_quote_ids")
    def _get_total_price(self):
        for record in self:
            approved_quotes = record.job_quote_ids.filtered(lambda e: e._is_approved())
            record.total_price = sum(approved_quotes.mapped("labor_cost"))

    @api.constrains("maintenance_scope_id")
    def restrict_only_one_unconfirmed_report_in_maintenance_scope(self):
        for record in self:
            report_ids = record.maintenance_scope_id.maintenance_scope_report_ids
            unconfirmed_reports = report_ids.filtered(lambda rp: not rp.finished_at)
            report_len = len(unconfirmed_reports)

            message = "Vui lòng hoàn thành báo cáo sửa chữa, trước khi tạo mới báo cáo cho hạng mục này!"
            if report_len >= 2:
                raise ValidationError(message)

    @api.constrains("maintenance_scope_id", "job_quote_ids")
    def _check_if_maintenance_scope_have_this_jobs(self):
        message = "Công việc của báo giá, và báo cáo sửa chữa không liên quan đến nhau!"
        for record in self:
            if record.maintenance_scope_id and record.job_quote_ids:
                scope_job_ids = record.maintenance_scope_id.job_ids
                quote_job_ids = record.job_quote_ids.mapped("job_id")

                have_different_jobs = not all(
                    [job_id in scope_job_ids for job_id in quote_job_ids]
                )

                if have_different_jobs:
                    raise ValidationError(message)

    @api.model_create_multi
    def create(self, vals_list):
        result = super(MaintenanceScopeReportTemplate, self).create(vals_list)

        for record in result:
            record._create_job_quotes_if_not_have()
            record._compute_week_number()
            record._set_deadline_based_on_expected_delivery_date()

        return result

    def write(self, vals):
        self.ensure_one()
        result = super(MaintenanceScopeReportTemplate, self).write(vals)

        if "expected_implement_date" in vals:
            self._compute_week_number()

        if "expected_delivery_date" in vals:
            self._set_deadline_based_on_expected_delivery_date()

        return result

    def unlink(self):
        for record in self:
            record.job_quote_ids.unlink()
        result = super(MaintenanceScopeReportTemplate, self).unlink()
        return result

    def action_propose_all_quotes(self):
        self.ensure_one()
        for quote in self.job_quote_ids:
            quote.action_propose()

    def action_unpropose_all_quotes(self):
        self.ensure_one()
        for quote in self.job_quote_ids:
            quote.action_unpropose()

    def action_approve_all_quotes(self):
        self.ensure_one()
        for quote in self.job_quote_ids:
            quote.action_approve()

    def action_reject_all_quotes(self):
        self.ensure_one()
        for quote in self.job_quote_ids:
            quote.action_reject()

    def _create_job_quotes_if_not_have(self):
        self.ensure_one()
        if not self.job_quote_ids:
            self._create_job_quotes()

    def _create_job_quotes(self):
        self.ensure_one()
        next_maintenance_date = self.maintenance_scope_id.next_maintenance_date

        for job_id in self.maintenance_scope_id.job_ids:
            self.job_quote_ids.create(
                {
                    "maintenance_scope_report_id": self.id,
                    "job_id": job_id.id,
                    "implement_date": next_maintenance_date,
                    "expected_delivery_date": self.expected_delivery_date,
                    "deadline": self.deadline,
                }
            )

    def _remove_job_quotes(self):
        self.ensure_one()
        job_quotes = self.job_quote_ids

        if job_quotes:
            job_quotes.unlink()

    def confirm_all_job_quotes(self):
        self.ensure_one()
        for quote in self.job_quote_ids:
            if quote.job_state == CONST.COMPLETED:
                quote._confirm_proposal()

    def unlink_and_create_job_quotes(self):
        self.ensure_one()
        self._remove_job_quotes()
        self._create_job_quotes_if_not_have()

    def _is_approved_report(self):
        self.ensure_one()
        return all([quote._is_approved() for quote in self.job_quote_ids])

    def finish_report_if_all_job_quote_finished(self):
        for record in self:
            if record._are_confirmed_all_job_quotes():
                self.maintenance_scope_id.last_maintenance_date = self.finished_at
                self.maintenance_scope_id.create_new_report_if_threshold_maintenance()

    def _are_confirmed_all_job_quotes(self):
        self.ensure_one()
        job_state_arr = self.job_quote_ids.mapped("job_state")
        all_confirmed = all([state == CONST.CONFIRMED for state in job_state_arr])

        return all_confirmed

    def _compute_week_number(self):
        for record in self:
            if record.expected_implement_date:
                week_number = record._get_week_number(record.expected_implement_date)
                record.week_number = week_number
            else:
                not_valid_week_number = 100
                record.week_number = not_valid_week_number

    def _set_deadline_based_on_expected_delivery_date(self):
        self.ensure_one()
        if self.expected_delivery_date:
            deadline_days = timedelta(days=self._get_deadline_days())
            deadline = self.expected_delivery_date - deadline_days
            self.deadline = deadline

    def _get_deadline_days(self):
        default_value_model = self._get_default_value_model()
        variable_name = (
            CONST_UTILITIES.INTEGER_SHIP_MAINTENANCE_SCOPE_REPORT_TEMPLATE_DEADLINE_DAY_COUNT
        )
        return default_value_model._get_default_value_by_variable_name(variable_name)

    def _get_default_value_model(self):
        model_name = "utilities.default.value"
        default_value_model = self.env[model_name].search([])

        return default_value_model

    def is_all_quotes_approved(self):
        self.ensure_one()
        approval_statuses = self.job_quote_ids.mapped("approval_status")

        return all(status == CONST.APPROVED for status in approval_statuses)

    def get_group_names_of_unapproved_quotes(self):
        self.ensure_one()
        unapproved_quotes = self.job_quote_ids.filtered(
            lambda e: not e._is_approved() and e.xml_id != CONST.EMPTY
        )
        xml_ids = unapproved_quotes.mapped("xml_id")
        unique_xml_ids = list(set(xml_ids))

        groups = [
            self.get_group_by_xml_id(xml_id)
            for xml_id in unique_xml_ids
            if isinstance(xml_id, str)
        ]
        group_names = [group.name for group in groups]

        return group_names

    def get_group_by_xml_id(self, xml_id):
        self.ensure_one()

        return self.env.ref(xml_id, raise_if_not_found=False)


class MaintenanceScopeReport(models.Model):
    _name = "ship.maintenance.scope.report"
    _description = "Maintenance scope report records"
    _inherit = ["ship.maintenance.scope.report.template"]
    _check_company_auto = True

    # relation field
    equipment_name = fields.Char(
        related="maintenance_scope_id.equipment_id.name",
        string="Equipment",
        store=False,
    )

    maintenance_type = fields.Selection(
        CONST.MAINTENANCE_TYPE,
        string="Maintenance Type",
        related="maintenance_scope_id.maintenance_type",
        tracking=True,
    )

    department_in_charge = fields.Selection(
        CONST.DEPARTMENT_IN_CHARGE,
        related="maintenance_scope_id.department_in_charge",
        string="Department in charge",
        tracking=True,
    )
    inspection_date = fields.Date(
        "Inspection Date",
        tracking=True,
    )

    _order = "expected_implement_date ASC"

    # company
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company
    )

    # sequential
    ref = fields.Char(string="Code", default=lambda self: _("New"))

    # relations
    maintenance_scope_id = fields.Many2one(
        "ship.maintenance.scope",
        string="Maintenance Scope",
        tracking=True,
    )
    job_quote_ids = fields.One2many(
        "ship.job.quote",
        "maintenance_scope_report_id",
        string="Job quote",
        tracking=True,
    )

    technical_incident_id = fields.Many2one("legis.technical.incident", readonly=True)
    serious_accident_team_id = fields.Many2one("legis.serious.accident.team", readonly=True)

    @api.constrains("technical_incident_id", "serious_accident_team_id")
    def only_technical_incident_or_serious_accident(self):
        for record in self:
            if record.technical_incident_id and record.serious_accident_team_id:
                message = "Chỉ được giao vật tư cho 1 model, liên hệ kỹ thuật!"
                raise ValidationError(message)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            vals["ref"] = self.env["ir.sequence"].next_by_code(
                "ship.maintenance.scope.report"
            )
        result = super(MaintenanceScopeReport, self).create(vals_list)

        for record in result:
            record.not_allow_perform_if_report_under_maintenance_at_docking()

        return result

    def write(self, vals):
        self.ensure_one()

        result = super(MaintenanceScopeReport, self).write(vals)

        self.not_allow_perform_if_report_under_maintenance_at_docking()
        return result

    def name_get(self):
        result = []
        for report in self:
            maintenance_scope_name = report.maintenance_scope_id.name or ""
            name = f"{maintenance_scope_name or _('New')}(Báo cáo)"
            result.append((report.id, name))
        return result

    def set_waiting_for_docking_state(self):
        self.ensure_one()
        context = dict(self.env.context)
        context.update({"allow_edit_under_maintenance": True})

        self.with_context(context).ref_job_state = CONST.WAITING_FOR_DOCKING

    def not_allow_perform_if_report_under_maintenance_at_docking(self):
        self.ensure_one()
        is_allow_edit = self.env.context.get("allow_edit_under_maintenance")
        scope_id = self.maintenance_scope_id

        if scope_id.is_docking and not is_allow_edit:
            message = "Hạng mục đang sửa chữa ở docking, vui lòng thao tác sau!"
            raise ValidationError(message)

    def generate_report(self):
        self.ensure_one()

        # Execute the SQL query
        query = """
        SELECT
        ms.id AS maintenance_scope_id,
        ms.name AS maintenance_scope_name,
        ms.description AS maintenance_scope_description,
        ms.department_in_charge AS maintenance_scope_department,
        ms.maintenance_type AS maintenance_scope_maintenance_type,
        MAX(ms.last_maintenance_date) AS maintenance_scope_last_maintenance_date,  
        ms.interval_days AS maintenance_scope_maintenance_interval_days,
        ms.utilization_time AS maintenance_scope_utilization_time,
        ms.allowed_usage_time AS maintenance_scope_allowed_usage_time,
        ms.equipment_id AS maintenance_scope_equipment_id,
        ms.ref AS maintenance_scope_ref,
        eq.id AS equipment_id,
        eq.name AS equipment_name,
        eq.description AS equipment_description,
        eq.calculate_consumption_by AS equipment_calculate_consumption_by,
        eq.last_recording_date AS equipment_last_recording_date,
        j.id AS job_id,
        j.name AS job_name,
        j.description AS job_description,
        j.assigned_group AS job_assigned_group,
        msr.finished_at as finished_date
    FROM
        ship_maintenance_scope ms
    JOIN
        ship_equipment eq ON ms.equipment_id = eq.id
    JOIN
        ship_job j ON ms.id = j.maintenance_scope_id
    JOIN
        ship_maintenance_scope_report msr ON ms.id = msr.maintenance_scope_id
    WHERE
        ms.maintenance_type = 'THRESHOLD'
        and 
        ms.interval_days is not null
        AND EXTRACT(YEAR FROM ms.last_maintenance_date) > 2022
    GROUP BY
        ms.id, ms.name, ms.description, ms.department_in_charge, ms.maintenance_type,
        ms.interval_days, ms.utilization_time, ms.allowed_usage_time,
        ms.equipment_id, ms.ref, eq.id, eq.name, eq.description,
        eq.calculate_consumption_by, eq.last_recording_date,
        j.id, j.name, j.description, j.assigned_group,
        msr.finished_at
    ORDER BY
        maintenance_scope_last_maintenance_date, maintenance_scope_name, job_name;
        """
        self.env.cr.execute(query)
        results = self.env.cr.fetchall()

        data = []
        for result in results:
            # Process the query results and construct the report data
            data.append(
                {
                    # Map the fields from the query result to the fields in the report
                    # Adjust field mappings as per your model definition
                    "maintenance_scope_id": result[0],
                    "maintenance_scope_name": result[1],
                    "maintenance_scope_description": result[2],
                    "maintenance_scope_department": result[3],
                    "maintenance_scope_maintenance_type": result[4],
                    "maintenance_scope_last_maintenance_date": result[5],
                    "maintenance_scope_maintenance_interval_days": result[6],
                    "maintenance_scope_utilization_time": result[7],
                    "maintenance_scope_allowed_usage_time": result[8],
                    "maintenance_scope_equipment_id": result[9],
                    "maintenance_scope_ref": result[10],
                    "equipment_id": result[11],
                    "equipment_name": result[12],
                    "equipment_description": result[13],
                    "equipment_calculate_consumption_by": result[14],
                    "equipment_last_recording_date": result[15],
                    "job_id": result[16],
                    "job_name": result[17],
                    "job_description": result[18],
                    "job_assigned_group": result[19],
                    "finished_date": result[20],
                    # Add other fields as needed
                }
            )

        return data

    def custom_export_pms_form_to_xlsx(self):
        # Function to adjust merged cells after inserting rows

        # Load the custom XLSX template
        template_path = (
            "/mnt/extra-addons/report_template/pms.xlsx"  # Replace with the actual path
        )
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###

        align_center = Alignment(horizontal="center", vertical="center")
        side_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
        )
        # date_object = fields.Date.from_string(self.proposed_date)
        # day = date_object.day
        # month = date_object.month
        # year = date_object.year
        ### Tên tàu
        worksheet_name = ["Boong", "Máy"]
        for ws in worksheet_name:
            worksheet = workbook[ws]
            worksheet.cell(
                row=3, column=2, value=f"M/V (Tên tàu): {self.company_id.name}"
            )
            worksheet[f"B3"].alignment = align_center

        ####PMS
        pms_data_machinery = []  # Initialize lists for different departments
        pms_data_boong = []

        pms_data = self.generate_report()
        pms_data_df = pd.DataFrame(pms_data)
        aggregated_data = (
            pms_data_df.groupby("maintenance_scope_id")["finished_date"]
            .apply(list)
            .reset_index()
        )
        pms_data_df.drop(columns=["finished_date"], inplace=True)
        pms_data_df.drop_duplicates(inplace=True)

        for index, maintainance_data in pms_data_df.iterrows():
            if maintainance_data["maintenance_scope_department"] == "MACHINERY":
                pms_data_machinery.append(maintainance_data)
            elif maintainance_data["maintenance_scope_department"] == "BOONG":
                pms_data_boong.append(maintainance_data)

        for pms_data_dep, department_name in [
            (pms_data_machinery, "Máy"),
            (pms_data_boong, "Boong"),
        ]:
            row = 6
            index = 0
            worksheet = workbook[department_name]
            for maintainance_data in pms_data_dep:
                worksheet.insert_rows(row)
                index += 1
                worksheet.cell(row=row, column=1, value=index)
                worksheet.cell(
                    row=row,
                    column=2,
                    value=maintainance_data["equipment_name"]
                    + "-"
                    + maintainance_data["maintenance_scope_name"],
                )
                worksheet.cell(
                    row=row,
                    column=3,
                    value=maintainance_data["job_name"],
                )
                worksheet.cell(
                    row=row,
                    column=4,
                    value=f"{int(int(maintainance_data['maintenance_scope_maintenance_interval_days'])/30)} tháng",
                )

                date_object = fields.Date.from_string(
                    maintainance_data["maintenance_scope_last_maintenance_date"]
                )
                day = date_object.day
                month = date_object.month
                year = date_object.year

                worksheet.cell(
                    row=row,
                    column=5,
                    value=f"{maintainance_data['maintenance_scope_last_maintenance_date']}",
                )

                date_object = fields.Date.from_string(
                    maintainance_data["maintenance_scope_last_maintenance_date"]
                )
                day = date_object.day
                month = date_object.month
                year = date_object.year
                end_of_year = datetime(2024, 12, 31)
                # Create a datetime object from the extracted components
                last_maintenance_date = datetime(year, month, day)
                current_date = last_maintenance_date
                expected_dates = []
                while current_date < end_of_year:
                    expected_dates.append(current_date)
                    current_date += timedelta(
                        days=int(
                            maintainance_data[
                                "maintenance_scope_maintenance_interval_days"
                            ]
                        )
                    )
                for date in expected_dates:
                    month_number = date.month
                    year = date.year
                    if year == datetime.now().year:
                        cell = worksheet.cell(
                            row=row,
                            column=month_number + 6,
                            value="Y",
                        )
                row += 1

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            border_style = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            worksheet_name = ["Boong", "Máy"]
            for ws in worksheet_name:
                worksheet = workbook[ws]
                for row in range(6, index + 1):
                    for col in range(1, 19):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = border_style

            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Kế hoạch PMS.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action

    def custom_export_pms_review_to_xlsx(self):
        # Load the custom XLSX template
        template_path = (
            "/mnt/extra-addons/report_template/pms.xlsx"  # Replace with the actual path
        )
        try:
            workbook = load_workbook(template_path)
            worksheet = workbook.active
        except Exception as e:
            raise UserError(f"Error loading the template: {str(e)}")

        ###

        align_center = Alignment(horizontal="center", vertical="center")
        side_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
        )
        # date_object = fields.Date.from_string(self.proposed_date)
        # day = date_object.day
        # month = date_object.month
        # year = date_object.year
        ### Tên tàu
        worksheet_name = ["Boong", "Máy"]
        for ws in worksheet_name:
            worksheet = workbook[ws]
            worksheet.cell(
                row=3, column=2, value=f"M/V (Tên tàu): {self.company_id.name}"
            )
            worksheet[f"B3"].alignment = align_center

        ####PMS
        pms_data_machinery = []  # Initialize lists for different departments
        pms_data_boong = []

        pms_data = self.generate_report()
        pms_data_df = pd.DataFrame(pms_data)
        aggregated_data = (
            pms_data_df.groupby("maintenance_scope_id")["finished_date"]
            .apply(lambda x: list(filter(None, x)))
            .reset_index()
        )
        pms_data_df.drop(columns=["finished_date"], inplace=True)
        pms_data_df.drop_duplicates(inplace=True)
        merged_data = pd.merge(
            pms_data_df, aggregated_data, on="maintenance_scope_id", how="left"
        )

        for index, maintainance_data in merged_data.iterrows():
            if maintainance_data["maintenance_scope_department"] == "MACHINERY":
                pms_data_machinery.append(maintainance_data)
            elif maintainance_data["maintenance_scope_department"] == "BOONG":
                pms_data_boong.append(maintainance_data)

        for pms_data_dep, department_name in [
            (pms_data_machinery, "Máy"),
            (pms_data_boong, "Boong"),
        ]:
            row = 6
            index = 0
            worksheet = workbook[department_name]
            for maintainance_data in pms_data_dep:
                worksheet.insert_rows(row)
                index += 1
                worksheet.cell(row=row, column=1, value=index)
                worksheet.cell(
                    row=row,
                    column=2,
                    value=maintainance_data["equipment_name"]
                    + "-"
                    + maintainance_data["maintenance_scope_name"],
                )
                worksheet.cell(
                    row=row,
                    column=3,
                    value=maintainance_data["job_name"],
                )
                worksheet.cell(
                    row=row,
                    column=4,
                    value=f"{int(int(maintainance_data['maintenance_scope_maintenance_interval_days'])/30)} tháng",
                )

                date_object = fields.Date.from_string(
                    maintainance_data["maintenance_scope_last_maintenance_date"]
                )
                day = date_object.day
                month = date_object.month
                year = date_object.year

                worksheet.cell(
                    row=row,
                    column=5,
                    value=f"{maintainance_data['maintenance_scope_last_maintenance_date']}",
                )

                date_object = fields.Date.from_string(
                    maintainance_data["maintenance_scope_last_maintenance_date"]
                )
                day = date_object.day
                month = date_object.month
                year = date_object.year
                end_of_year = datetime(2024, 12, 31)
                # Create a datetime object from the extracted components
                last_maintenance_date = datetime(year, month, day)
                current_date = last_maintenance_date
                expected_dates = []
                while current_date < end_of_year:
                    expected_dates.append(current_date)
                    current_date += timedelta(
                        days=int(
                            maintainance_data[
                                "maintenance_scope_maintenance_interval_days"
                            ]
                        )
                    )
                for date in expected_dates:
                    month_number = date.month
                    year = date.year
                    if year == datetime.now().year:
                        cell = worksheet.cell(
                            row=row,
                            column=month_number + 6,
                            value="Y",
                        )
                for date in maintainance_data["finished_date"]:
                    month_number = date.month
                    year = date.year
                    if year == datetime.now().year:
                        cell = worksheet.cell(
                            row=row,
                            column=month_number + 6,
                            value=date,
                        )
                row += 1

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            # Save the modified XLSX file to the temporary file
            border_style = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            worksheet_name = ["Boong", "Máy"]
            for ws in worksheet_name:
                worksheet = workbook[ws]
                for row in range(6, index + 1):
                    for col in range(1, 19):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = border_style

            workbook.save(temp_file.name)

            # Load the saved XLSX file for further processing
            temp_workbook = load_workbook(temp_file.name)
            temp_worksheet = temp_workbook.active

            # Save the modified workbook back to the temporary file
            temp_workbook.save(temp_file.name)

        # Read the temporary XLSX file as binary data
        with open(temp_file.name, "rb") as file:
            binary_data = file.read()

        # Create an attachment
        filename = f"Review PMS.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(binary_data),
                "res_model": self._name,
                "res_id": self.id,
                "public": True,
                "res_field": "custom_export",  # Replace with the appropriate field name
            }
        )

        action = {
            "name": filename,
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}/{filename}",
            "target": "self",
        }
        # Return an action to open the attachment
        return action

    def action_is_docking(self):
        self.ensure_one()
        self.maintenance_scope_id.is_docking = True
